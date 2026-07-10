import json
import tempfile
import threading
import time
import unittest
import urllib.request
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from taotian_price_tool.bi import BiImportResult, BiTemplateItem
from taotian_price_tool.models import ProductSnapshot
from taotian_price_tool import webapp as webapp_module
from taotian_price_tool.webapp import DashboardRequestHandler, DashboardServer, TaskSession, WebAppController


class FakeBrowserManager:
    def __init__(self):
        self.opened = False
        self.close_calls = 0

    def open(self):
        self.opened = True
        return {"status": "opened", "message": "opened"}

    def close(self, timeout=3.0):
        self.close_calls += 1
        self.opened = False

    def status(self):
        return {"running": self.opened, "last_error": "", "current_url": ""}


class StubbornBrowserManager(FakeBrowserManager):
    def close(self, timeout=3.0):
        self.close_calls += 1
        self.opened = True


class FakeScraper:
    def __init__(self, snapshots=None):
        self.calls = []
        self.snapshots = list(snapshots or [])

    def fetch(self, url):
        self.calls.append(url)
        if self.snapshots:
            return self.snapshots.pop(0)
        return ProductSnapshot(
            status="success",
            title="Redmi Buds 6 入耳式降噪耳机",
            price="159",
            detail_text="Redmi Buds 6 入耳式降噪耳机 ￥159",
        )


class FakeBiClient:
    def __init__(self, items):
        self.items = items
        self.configs = []

    def import_goods(self, config, *, template_brands=None):
        self.configs.append(config)
        return BiImportResult(
            import_count=len(self.items),
            earphone_count=len(self.items),
            items=list(self.items),
            message=f"识别到 {len(self.items)} 个耳机新品",
        )


class FakeFeishuClient:
    def __init__(self, app_id="cli_test", app_secret="secret"):
        self.app_id = app_id
        self.app_secret = app_secret
        self.fields_requested = []
        self.created_records = []
        self.existing_records = []
        self.fail_list_records = False
        self.list_records_calls = 0

    def list_spreadsheets(self):
        return ([{"name": "汇总表"}, {"name": "上新sop"}], None)

    def list_tables(self, app_token):
        return [
            {"id": "tbl_summary", "name": "汇总表"},
            {"id": "tbl_sop", "name": "上新sop"},
        ]

    def list_fields(self, app_token, table_id):
        self.fields_requested.append(table_id)
        return [
            {"name": "品牌", "type": 1},
            {"name": "链接", "type": 1},
            {"name": "形态", "type": 1},
            {"name": "价格", "type": 1},
            {"name": "上架日期", "type": 1},
            {"name": "标题", "type": 1},
        ]

    def list_records(self, app_token, table_id):
        self.list_records_calls += 1
        if self.fail_list_records:
            raise RuntimeError("records unavailable")
        return list(self.existing_records)

    def batch_create_records(self, app_token, table_id, records, field_types):
        self.created_records.extend(records)
        return len(records)


def workbook_bytes(links=None) -> bytes:
    links = links or ["https://detail.tmall.com/item.htm?id=1"]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = Path(tmp.name)
    wb = Workbook()
    ws = wb.active
    for index, link in enumerate(links, start=3):
        ws.cell(row=index, column=3).value = link
    wb.save(path)
    data = path.read_bytes()
    path.unlink()
    return data


def template_workbook_bytes() -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = Path(tmp.name)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "品牌"
    ws["B1"] = "上架日期"
    ws["C1"] = "型号"
    ws["D1"] = "形态"
    ws["E1"] = "价格"
    ws.merge_cells("A4:A6")
    ws["A4"] = "索爱"
    ws.merge_cells("B4:K6")
    ws["B4"] = "无上新"
    ws.merge_cells("A7:A8")
    ws["A7"] = "小米"
    ws.merge_cells("B7:K8")
    ws["B7"] = "旧内容"
    wb.save(path)
    data = path.read_bytes()
    path.unlink()
    return data


def new_template_workbook_bytes() -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = Path(tmp.name)
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("B1:F1")
    ws["B1"] = "上新监控"
    headers = ["品牌", "上架日期", "链接", "标题", "形态", "价格"]
    for column, header in enumerate(headers, start=1):
        ws.cell(row=2, column=column).value = header
    ws.append(["索爱", None, None, None, None, None])
    ws.append(["小米", None, None, None, None, None])
    wb.save(path)
    data = path.read_bytes()
    path.unlink()
    return data


class WebAppControllerTests(unittest.TestCase):
    def test_upload_saves_excel_copy_and_updates_status(self):
        with tempfile.TemporaryDirectory() as tmp:
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )

            result = controller.save_upload("input.xlsx", workbook_bytes())
            status = controller.status()

            self.assertEqual("input.xlsx", result["file_name"])
            self.assertEqual(1, result["link_count"])
            self.assertTrue(Path(result["path"]).exists())
            self.assertEqual("input.xlsx", status["selected_file_name"])

    def test_start_collection_reuses_open_taobao_browser_and_writes_output(self):
        with tempfile.TemporaryDirectory() as tmp:
            browser = FakeBrowserManager()
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=browser,
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())
            browser.open()

            start = controller.start_collection()
            self._wait_for_state(controller, "completed")
            status = controller.status()

            self.assertEqual("started", start["status"])
            self.assertEqual(0, browser.close_calls)
            self.assertTrue(browser.opened)
            self.assertEqual("completed", status["run_state"])
            self.assertTrue(Path(status["output_path"]).exists())

    def test_start_collection_does_not_require_closing_open_taobao_browser(self):
        with tempfile.TemporaryDirectory() as tmp:
            browser = StubbornBrowserManager()
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=browser,
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())
            browser.open()

            controller.start_collection()
            self._wait_for_state(controller, "completed")

            self.assertEqual(0, browser.close_calls)
            self.assertEqual("completed", controller.status()["run_state"])

    def test_verification_waits_for_manual_action_then_resumes_current_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="needs_verification", note="页面进入登录/身份验证，需要人工处理"),
                    ProductSnapshot(status="success", title="Redmi Buds 入耳式耳机", price="159"),
                    ProductSnapshot(status="success", title="AirPods 半入耳耳机", price="199"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload(
                "input.xlsx",
                workbook_bytes(
                    [
                        "https://detail.tmall.com/item.htm?id=1",
                        "https://detail.tmall.com/item.htm?id=2",
                    ]
                ),
            )

            controller.start_collection()
            self._wait_for_state(controller, "waiting_manual_verification")
            paused = controller.status()

            self.assertEqual(3, paused["current_row"])
            self.assertIn("id=1", paused["current_url"])
            self.assertTrue(paused["can_resume"])
            self.assertEqual("请在已打开的淘宝浏览器完成登录/验证后继续当前行", paused["next_action"])
            self.assertEqual(["https://detail.tmall.com/item.htm?id=1"], scraper.calls)

            controller.resume_current_row()
            self._wait_for_state(controller, "completed")
            final = controller.status()

            self.assertEqual(
                [
                    "https://detail.tmall.com/item.htm?id=1",
                    "https://detail.tmall.com/item.htm?id=1",
                    "https://detail.tmall.com/item.htm?id=2",
                ],
                scraper.calls,
            )
            self.assertEqual(2, final["completed"])
            self.assertEqual(0, final["failed"])

    def test_resume_current_row_does_not_close_open_taobao_browser(self):
        with tempfile.TemporaryDirectory() as tmp:
            browser = FakeBrowserManager()
            browser.open()
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="needs_verification", note="页面进入登录/身份验证，需要人工处理"),
                    ProductSnapshot(status="success", title="Redmi Buds 入耳式耳机", price="159"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=browser,
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())

            controller.start_collection()
            self._wait_for_state(controller, "waiting_manual_verification")
            controller.resume_current_row()
            self._wait_for_state(controller, "completed")

            self.assertEqual(0, browser.close_calls)
            self.assertTrue(browser.opened)

    def test_off_shelf_row_is_processed_without_shape_and_continues(self):
        with tempfile.TemporaryDirectory() as tmp:
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="off_shelf", note="详情页提示商品已下架"),
                    ProductSnapshot(status="success", title="AirPods 半入耳耳机", price="199"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload(
                "input.xlsx",
                workbook_bytes(
                    [
                        "https://detail.tmall.com/item.htm?id=off",
                        "https://detail.tmall.com/item.htm?id=ok",
                    ]
                ),
            )

            controller.start_collection()
            self._wait_for_state(controller, "completed")
            status = controller.status()
            copied = load_workbook(status["output_path"]).active

            self.assertEqual("completed", status["run_state"])
            self.assertEqual(2, status["completed"])
            self.assertEqual(0, status["failed"])
            self.assertEqual("", copied["D3"].value or "")
            self.assertEqual("", copied["E3"].value or "")
            self.assertTrue(any(row["row"] == 3 and row["status"] == "off_shelf" for row in status["logs"]))

    def test_repeated_verification_updates_current_log_instead_of_spamming(self):
        with tempfile.TemporaryDirectory() as tmp:
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="needs_verification", note="第一次验证"),
                    ProductSnapshot(status="needs_verification", note="第二次验证"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())

            controller.start_collection()
            self._wait_for_state(controller, "waiting_manual_verification")
            controller.resume_current_row()
            self._wait_for_calls(scraper, 2)
            status = controller.status()
            verification_logs = [
                row for row in status["logs"] if row.get("row") == 3 and row.get("status") == "needs_verification"
            ]

            self.assertEqual(1, len(verification_logs))
            self.assertEqual("第二次验证", verification_logs[0]["note"])

    def test_skip_current_row_marks_failure_and_continues_next_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="needs_verification", note="页面进入登录/身份验证，需要人工处理"),
                    ProductSnapshot(status="success", title="AirPods 半入耳耳机", price="199"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload(
                "input.xlsx",
                workbook_bytes(
                    [
                        "https://detail.tmall.com/item.htm?id=1",
                        "https://detail.tmall.com/item.htm?id=2",
                    ]
                ),
            )

            controller.start_collection()
            self._wait_for_state(controller, "waiting_manual_verification")
            controller.skip_current_row()
            self._wait_for_state(controller, "completed")
            status = controller.status()

            self.assertEqual(
                [
                    "https://detail.tmall.com/item.htm?id=1",
                    "https://detail.tmall.com/item.htm?id=2",
                ],
                scraper.calls,
            )
            self.assertEqual(1, status["completed"])
            self.assertEqual(1, status["failed"])
            self.assertTrue(any(row["row"] == 3 and row["status"] == "failed" for row in status["logs"]))

    def test_bi_import_then_template_fill_reuses_collection_flow(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="索爱",
                shop_name="SOAIY旗舰店",
                goods_name="SOAIY 开放式耳夹耳机",
                goods_link="http://a.m.taobao.com/i106.htm?&sid=origin&jose=1",
                fetch_link="https://detail.tmall.com/item.htm?id=106",
                item_id="106",
                on_sale_time="2026-06-20",
            )
            scraper = FakeScraper(
                [ProductSnapshot(status="success", title="SOAIY 开放式耳夹耳机", price="159")]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                bi_client=FakeBiClient([bi_item]),
                sleep=lambda seconds: None,
            )
            controller.save_upload("template.xlsx", template_workbook_bytes())

            imported = controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})
            started = controller.start_bi_template_fill()
            self._wait_for_state(controller, "completed")
            status = controller.status()
            copied = load_workbook(status["output_path"]).active

            self.assertEqual(1, imported["bi_earphone_count"])
            self.assertEqual("started", started["status"])
            self.assertEqual(["https://detail.tmall.com/item.htm?id=106"], scraper.calls)
            self.assertEqual("6月20", copied["B4"].value)
            self.assertEqual("http://a.m.taobao.com/i106.htm?&sid=origin&jose=1", copied["C4"].value)
            self.assertEqual("耳夹式", copied["D4"].value)
            self.assertEqual("159", copied["E4"].value)
            self.assertEqual("无上新", copied["B7"].value)
            self.assertEqual(1, status["bi_earphone_count"])

    def test_bi_template_fill_writes_new_template_title_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="索爱",
                shop_name="SOAIY旗舰店",
                goods_name="SOAIY 开放式耳夹耳机",
                goods_link="https://detail.tmall.com/item.htm?id=106",
                fetch_link="https://detail.tmall.com/item.htm?id=106",
                item_id="106",
                on_sale_time="2026-06-20",
            )
            scraper = FakeScraper(
                [ProductSnapshot(status="success", title="淘宝详情耳夹式标题", price="159")]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                bi_client=FakeBiClient([bi_item]),
                sleep=lambda seconds: None,
            )
            controller.save_upload("new_template.xlsx", new_template_workbook_bytes())

            controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})
            controller.start_bi_template_fill()
            self._wait_for_state(controller, "completed")
            status = controller.status()
            copied = load_workbook(status["output_path"]).active

            self.assertEqual("6月20号", copied["B3"].value)
            self.assertEqual("https://detail.tmall.com/item.htm?id=106", copied["C3"].value)
            self.assertEqual("SOAIY 开放式耳夹耳机", copied["D3"].value)
            self.assertEqual("耳夹式", copied["E3"].value)
            self.assertEqual("159", copied["F3"].value)
            self.assertEqual([None, None, None, None, None], [copied.cell(4, col).value for col in range(2, 7)])
            self.assertEqual(["https://detail.tmall.com/item.htm?id=106"], scraper.calls)

    def test_bi_template_fill_writes_same_link_with_different_dates_as_two_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            items = [
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY Auro Ace蓝牙耳机小头戴式",
                    goods_link="https://detail.tmall.com/item.htm?id=106",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-18",
                ),
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY Auro Ace蓝牙耳机小头戴式",
                    goods_link="https://detail.tmall.com/item.htm?id=106",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-14",
                ),
            ]
            scraper = FakeScraper(
                [ProductSnapshot(status="success", title="SOAIY Auro Ace蓝牙耳机小头戴式", price="299")]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                bi_client=FakeBiClient(items),
                sleep=lambda seconds: None,
            )
            controller.save_upload("new_template.xlsx", new_template_workbook_bytes())

            imported = controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})
            controller.start_bi_template_fill()
            self._wait_for_state(controller, "completed")
            status = controller.status()
            copied = load_workbook(status["output_path"]).active

            self.assertEqual(2, imported["bi_earphone_count"])
            self.assertEqual(["6月18号", "6月14号"], [copied["B3"].value, copied["B4"].value])
            self.assertEqual(
                ["https://detail.tmall.com/item.htm?id=106", "https://detail.tmall.com/item.htm?id=106"],
                [copied["C3"].value, copied["C4"].value],
            )
            self.assertEqual(["SOAIY Auro Ace蓝牙耳机小头戴式", "SOAIY Auro Ace蓝牙耳机小头戴式"], [copied["D3"].value, copied["D4"].value])
            self.assertEqual(["299", "299"], [copied["F3"].value, copied["F4"].value])
            self.assertEqual(["https://detail.tmall.com/item.htm?id=106"], scraper.calls)

    def test_bi_template_fill_skips_off_shelf_goods(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="索爱",
                shop_name="SOAIY旗舰店",
                goods_name="SOAIY 耳夹式耳机",
                goods_link="https://detail.tmall.com/item.htm?id=106",
                fetch_link="https://detail.tmall.com/item.htm?id=106",
                item_id="106",
                on_sale_time="2026-06-20",
            )
            scraper = FakeScraper([ProductSnapshot(status="off_shelf", note="详情页提示商品已下架")])
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                bi_client=FakeBiClient([bi_item]),
                sleep=lambda seconds: None,
            )
            controller.save_upload("new_template.xlsx", new_template_workbook_bytes())

            controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})
            controller.start_bi_template_fill()
            self._wait_for_state(controller, "completed")
            status = controller.status()
            copied = load_workbook(status["output_path"]).active

            self.assertEqual([None, None, None, None, None], [copied.cell(3, col).value for col in range(2, 7)])
            self.assertTrue(any("已跳过" in row.get("note", "") for row in status["logs"]))

    def test_custom_output_directory_is_used_for_result_copy(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp) / "chosen-output"
            output_dir.mkdir()
            controller = WebAppController(
                base_dir=Path(tmp) / "app",
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())

            controller.save_settings({"output_dir": str(output_dir)})
            controller.start_collection()
            self._wait_for_state(controller, "completed")
            status = controller.status()

            self.assertEqual(str(output_dir), status["output_dir"])
            self.assertEqual("自定义", status["output_dir_source"])
            self.assertEqual(output_dir, Path(status["output_path"]).parent)

    def test_select_output_dir_persists_in_local_settings(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_dir = Path(tmp) / "selected"
            output_dir.mkdir()
            base_dir = Path(tmp) / "app"
            controller = WebAppController(
                base_dir=base_dir,
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                folder_picker=lambda: str(output_dir),
                sleep=lambda seconds: None,
            )

            selected = controller.select_output_dir()
            restarted = WebAppController(
                base_dir=base_dir,
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )

            self.assertEqual(str(output_dir), selected["output_dir"])
            self.assertEqual(str(output_dir), restarted.status()["output_dir"])

    def test_upload_template_does_not_clear_previous_bi_import_status(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds",
                goods_link="http://a.m.taobao.com/i1059057030771.htm?&sid=origin&jose=1",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                bi_client=FakeBiClient([bi_item]),
                sleep=lambda seconds: None,
            )

            controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})
            controller.save_upload("template.xlsx", template_workbook_bytes())
            status = controller.status()

            self.assertEqual(1, status["bi_import_count"])
            self.assertEqual(1, status["bi_earphone_count"])
            self.assertEqual("completed", status["bi_import_status"])

    def test_clear_bi_import_removes_stale_import_state_and_restores_excel_mode(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds",
                goods_link="https://detail.tmall.com/item.htm?id=1059057030771",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                bi_client=FakeBiClient([bi_item]),
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes(["https://detail.tmall.com/item.htm?id=1"]))
            controller.import_bi_goods({"days_type": 7, "end_date": "2026-06-20"})

            cleared = controller.clear_bi_import()
            status = controller.status()

            self.assertEqual("cleared", cleared["status"])
            self.assertEqual(0, status["bi_import_count"])
            self.assertEqual(0, status["bi_earphone_count"])
            self.assertEqual("idle", status["bi_import_status"])
            self.assertEqual("", status["bi_import_message"])
            self.assertEqual(1, status["link_count"])
            self.assertEqual("excel", status["active_mode"])
            self.assertFalse(status["can_start_bi_fill"])
            self.assertTrue(any(row.get("status") == "info" and "BI" in row.get("note", "") for row in status["logs"]))

    def test_resume_logs_browser_session_reuse(self):
        with tempfile.TemporaryDirectory() as tmp:
            scraper = FakeScraper(
                [
                    ProductSnapshot(status="needs_verification", note="页面进入登录/身份验证，需要人工处理"),
                    ProductSnapshot(status="success", title="Redmi Buds 入耳式耳机", price="159"),
                ]
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: scraper,
                sleep=lambda seconds: None,
            )
            controller.save_upload("input.xlsx", workbook_bytes())

            controller.start_collection()
            self._wait_for_state(controller, "waiting_manual_verification")
            controller.resume_current_row()
            self._wait_for_state(controller, "completed")

            notes = [row.get("note", "") for row in controller.status()["logs"]]
            self.assertTrue(any("复用淘宝浏览器会话" in note for note in notes))

    def test_feishu_selected_table_name_matches_selected_table(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller._feishu_client = feishu
            controller.state.feishu_connected = True

            result = controller._load_feishu_tables_and_fields("base_token", select_table_id="tbl_sop")
            status = controller.status()

            self.assertEqual("tbl_sop", status["feishu_table_id"])
            self.assertEqual("上新sop", status["feishu_table_name"])
            self.assertEqual(["tbl_sop"], feishu.fields_requested)
            self.assertIn("6/6", result["mapping_summary"])

    def test_feishu_restore_uses_saved_credentials_and_table(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            saved = {
                "app_id": "cli_saved",
                "app_secret": "secret_saved",
                "app_token": "https://example.feishu.cn/base/base_token",
                "table_id": "tbl_sop",
                "table_name": "上新sop",
            }
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.load_feishu_credentials", return_value=saved), patch(
                "taotian_price_tool.webapp.FeishuClient", return_value=feishu
            ):
                restored = controller.feishu_restore()

            status = controller.status()
            self.assertEqual("restored", restored["status"])
            self.assertEqual("tbl_sop", status["feishu_table_id"])
            self.assertEqual("上新sop", status["feishu_table_name"])
            self.assertEqual(["tbl_sop"], feishu.fields_requested)

    def test_feishu_connect_defaults_to_shangxin_sop_when_no_saved_table(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            saved_calls = []

            def capture_save(app_id, app_secret, app_token="", table_id="", table_name=""):
                saved_calls.append(
                    {
                        "app_id": app_id,
                        "app_secret": app_secret,
                        "app_token": app_token,
                        "table_id": table_id,
                        "table_name": table_name,
                    }
                )

            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.FeishuClient", return_value=feishu), patch(
                "taotian_price_tool.webapp.save_feishu_credentials", side_effect=capture_save
            ):
                result = controller.feishu_connect(
                    {"app_id": "cli_saved", "app_secret": "secret_saved", "app_token": "Q1oFabc123"}
                )

            status = controller.status()
            self.assertEqual("tbl_sop", result["table_id"])
            self.assertEqual("上新sop", result["table_name"])
            self.assertEqual("tbl_sop", status["feishu_table_id"])
            self.assertEqual("上新sop", status["feishu_table_name"])
            self.assertEqual(["tbl_sop"], feishu.fields_requested)
            self.assertEqual("tbl_sop", saved_calls[-1]["table_id"])

    def test_feishu_restore_falls_back_to_shangxin_sop_when_saved_table_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            saved = {
                "app_id": "cli_saved",
                "app_secret": "secret_saved",
                "app_token": "Q1oFabc123",
                "table_id": "tbl_missing",
                "table_name": "旧表",
            }
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.load_feishu_credentials", return_value=saved), patch(
                "taotian_price_tool.webapp.FeishuClient", return_value=feishu
            ), patch("taotian_price_tool.webapp.save_feishu_credentials"):
                restored = controller.feishu_restore()

            status = controller.status()
            self.assertEqual("restored", restored["status"])
            self.assertEqual("tbl_sop", status["feishu_table_id"])
            self.assertEqual("上新sop", status["feishu_table_name"])
            self.assertEqual(["tbl_sop"], feishu.fields_requested)

    def test_feishu_restore_reports_saved_secret_invalid(self):
        with tempfile.TemporaryDirectory() as tmp:
            saved = {
                "app_id": "cli_saved",
                "app_secret": "bad_secret",
                "app_token": "base_token",
                "table_id": "tbl_sop",
                "table_name": "上新sop",
            }
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.load_feishu_credentials", return_value=saved), patch(
                "taotian_price_tool.webapp.FeishuClient",
                side_effect=RuntimeError("{'code': 10014, 'msg': 'app secret invalid'}"),
            ):
                restored = controller.feishu_restore()

            self.assertFalse(restored["ok"])
            self.assertEqual("restore_failed", restored["status"])
            self.assertIn("保存的 App Secret 已失效", restored["error"])
            self.assertFalse(controller.status()["feishu_connected"])

    def test_feishu_connect_does_not_save_when_table_load_fails(self):
        class FailingTablesFeishu(FakeFeishuClient):
            def list_tables(self, app_token):
                raise RuntimeError("table load failed")

        with tempfile.TemporaryDirectory() as tmp:
            feishu = FailingTablesFeishu()
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.FeishuClient", return_value=feishu), patch(
                "taotian_price_tool.webapp.save_feishu_credentials"
            ) as save_mock:
                with self.assertRaises(RuntimeError):
                    controller.feishu_connect(
                        {"app_id": "cli_saved", "app_secret": "secret_saved", "app_token": "Q1oFabc123"}
                    )

            save_mock.assert_not_called()
            status = controller.status()
            self.assertFalse(status["feishu_connected"])
            self.assertEqual("", status["feishu_url_saved"])

    def test_feishu_load_fields_saves_manual_table_as_default(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            saved_calls = []

            def capture_save(app_id, app_secret, app_token="", table_id="", table_name=""):
                saved_calls.append({"table_id": table_id, "table_name": table_name})

            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller._feishu_client = feishu
            controller.state.feishu_connected = True
            controller.state.feishu_app_token = "Q1oFabc123"
            with patch("taotian_price_tool.webapp.save_feishu_credentials", side_effect=capture_save):
                result = controller.feishu_load_fields({"table_id": "tbl_summary"})

            status = controller.status()
            self.assertEqual("tbl_summary", result["table_id"])
            self.assertEqual("汇总表", result["table_name"])
            self.assertEqual("tbl_summary", status["feishu_table_id_saved"])
            self.assertEqual("汇总表", status["feishu_table_name_saved"])
            self.assertEqual({"table_id": "tbl_summary", "table_name": "汇总表"}, saved_calls[-1])

    def test_saved_placeholder_feishu_values_are_not_treated_as_connected(self):
        with tempfile.TemporaryDirectory() as tmp:
            saved = {
                "app_id": "cli_test",
                "app_secret": "secret_saved",
                "app_token": "base_token",
                "table_id": "tbl_sop",
                "table_name": "上新sop",
            }
            with patch("taotian_price_tool.webapp.load_feishu_credentials", return_value=saved), patch(
                "taotian_price_tool.webapp.FeishuClient", side_effect=AssertionError("placeholder should not connect")
            ):
                controller = WebAppController(
                    base_dir=Path(tmp),
                    browser_manager=FakeBrowserManager(),
                    scraper_factory=lambda: FakeScraper(),
                    sleep=lambda seconds: None,
                )
                status = controller.status()
                restored = controller.feishu_restore()

            self.assertFalse(status["feishu_connected"])
            self.assertEqual("", status["feishu_app_id_saved"])
            self.assertEqual("", status["feishu_url_saved"])
            self.assertFalse(restored["ok"])
            self.assertEqual("invalid", restored["status"])

    def test_stale_heartbeat_does_not_shutdown_active_collection(self):
        with tempfile.TemporaryDirectory() as tmp:
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller._last_heartbeat = 100.0

            controller.state.run_state = "running"
            self.assertFalse(
                webapp_module._should_shutdown_for_stale_heartbeat(controller, now=200.0, timeout=10.0)
            )

            controller.state.run_state = "waiting_manual_verification"
            self.assertFalse(
                webapp_module._should_shutdown_for_stale_heartbeat(controller, now=200.0, timeout=10.0)
            )

            controller.state.run_state = "completed"
            self.assertFalse(
                webapp_module._should_shutdown_for_stale_heartbeat(controller, now=200.0, timeout=10.0)
            )

            controller.state.run_state = "idle"
            self.assertTrue(
                webapp_module._should_shutdown_for_stale_heartbeat(controller, now=200.0, timeout=10.0)
            )

    def test_feishu_connect_sanitizes_concatenated_saved_and_pasted_values(self):
        with tempfile.TemporaryDirectory() as tmp:
            created = {}

            def make_feishu(app_id, app_secret):
                created["app_id"] = app_id
                created["app_secret"] = app_secret
                return FakeFeishuClient(app_id=app_id, app_secret=app_secret)

            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            with patch("taotian_price_tool.webapp.FeishuClient", side_effect=make_feishu):
                result = controller.feishu_connect(
                    {
                        "app_id": "cli_testcli_demo1234567890",
                        "app_secret": "secret",
                        "app_token": "base_tokenhttps://xcnbc7loouq4.feishu.cn/base/Q1oFabc123",
                    }
                )

            status = controller.status()
            self.assertTrue(result["ok"])
            self.assertEqual("cli_demo1234567890", created["app_id"])
            self.assertEqual("cli_demo1234567890", status["feishu_app_id_saved"])
            self.assertEqual("Q1oFabc123", status["feishu_app_token"])
            self.assertEqual("Q1oFabc123", status["feishu_url_saved"])

    def test_start_feishu_collection_preserves_ultra_conservative_intensity(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds 入耳式耳机",
                goods_link="https://detail.tmall.com/item.htm?id=1059057030771",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            controller._feishu_client = FakeFeishuClient()
            controller.state.feishu_connected = True
            controller.state.feishu_app_token = "base_token"
            controller.bi_items = [bi_item]
            controller.state.bi_earphone_count = 1
            controller.save_settings({"intensity": "超保守"})

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual("超保守", controller.status()["intensity"])

    def test_feishu_writes_existing_string_link_with_same_item_id_and_date(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds 入耳式耳机",
                goods_link="https://detail.tmall.com/item.htm?id=1059057030771",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            feishu = FakeFeishuClient()
            feishu.existing_records = [
                {
                    "品牌": "小米",
                    "链接": "http://a.m.taobao.com/i1059057030771.htm?&sid=origin&jose=1",
                    "上架日期": "2026-06-20",
                }
            ]
            controller = self._feishu_collection_controller(tmp, feishu, [bi_item])

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(1, len(feishu.created_records))
            self.assertEqual("小米", feishu.created_records[0]["品牌"])
            self.assertEqual(0, feishu.list_records_calls)
            self.assertFalse(any("重复" in row.get("note", "") for row in controller.status()["logs"]))

    def test_feishu_existing_duplicate_earphone_writes_product_row_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds 入耳式耳机",
                goods_link="https://detail.tmall.com/item.htm?id=1059057030771",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            feishu = FakeFeishuClient()
            feishu.existing_records = [
                {
                    "品牌": "小米",
                    "链接": "http://a.m.taobao.com/i1059057030771.htm?&sid=origin&jose=1",
                    "上架日期": "2026-06-20",
                }
            ]
            controller = self._feishu_collection_controller(tmp, feishu, [bi_item])
            controller.state.checked_brands = ["小米"]

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(1, len(feishu.created_records))
            self.assertIn("链接", feishu.created_records[0])
            self.assertEqual(0, feishu.list_records_calls)

    def test_feishu_writes_existing_url_object_link_with_same_item_id_and_date(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="索爱",
                shop_name="SOAIY旗舰店",
                goods_name="SOAIY 开放式耳夹耳机",
                goods_link="https://detail.tmall.com/item.htm?id=106",
                fetch_link="https://detail.tmall.com/item.htm?id=106",
                item_id="106",
                on_sale_time="2026-06-20",
            )
            feishu = FakeFeishuClient()
            feishu.existing_records = [
                {
                    "品牌": "索爱",
                    "链接": {"link": "http://a.m.taobao.com/i106.htm?&sid=origin", "text": "商品链接"},
                    "上架日期": "2026-06-20",
                }
            ]
            controller = self._feishu_collection_controller(tmp, feishu, [bi_item])

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(1, len(feishu.created_records))
            self.assertEqual("索爱", feishu.created_records[0]["品牌"])
            self.assertEqual(0, feishu.list_records_calls)

    def test_feishu_keeps_duplicate_items_inside_same_bi_batch(self):
        with tempfile.TemporaryDirectory() as tmp:
            items = [
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY 开放式耳夹耳机",
                    goods_link="http://a.m.taobao.com/i106.htm?&sid=origin",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-20",
                ),
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY 开放式耳夹耳机",
                    goods_link="https://detail.tmall.com/item.htm?id=106",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-20",
                ),
            ]
            feishu = FakeFeishuClient()
            controller = self._feishu_collection_controller(tmp, feishu, items)

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(2, len(feishu.created_records))

    def test_feishu_allows_same_item_on_different_sale_dates(self):
        with tempfile.TemporaryDirectory() as tmp:
            items = [
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY 开放式耳夹耳机",
                    goods_link="https://detail.tmall.com/item.htm?id=106",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-20",
                ),
                BiTemplateItem(
                    brand="索爱",
                    shop_name="SOAIY旗舰店",
                    goods_name="SOAIY 开放式耳夹耳机",
                    goods_link="https://detail.tmall.com/item.htm?id=106",
                    fetch_link="https://detail.tmall.com/item.htm?id=106",
                    item_id="106",
                    on_sale_time="2026-06-18",
                ),
            ]
            feishu = FakeFeishuClient()
            controller = self._feishu_collection_controller(tmp, feishu, items)

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(2, len(feishu.created_records))

    def test_feishu_keeps_existing_no_product_brand_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            feishu = FakeFeishuClient()
            feishu.existing_records = [{"品牌": "索爱"}]
            controller = self._feishu_collection_controller(tmp, feishu, [])
            controller.state.checked_brands = ["索爱"]

            records = controller._build_feishu_records(controller.session)

            self.assertEqual([{"品牌": "索爱"}], records)
            self.assertEqual(0, feishu.list_records_calls)

    def test_feishu_existing_record_read_failure_does_not_block_write(self):
        with tempfile.TemporaryDirectory() as tmp:
            bi_item = BiTemplateItem(
                brand="小米",
                shop_name="小米官方旗舰店",
                goods_name="Redmi Buds 入耳式耳机",
                goods_link="https://detail.tmall.com/item.htm?id=1059057030771",
                fetch_link="https://detail.tmall.com/item.htm?id=1059057030771",
                item_id="1059057030771",
                on_sale_time="2026-06-20",
            )
            feishu = FakeFeishuClient()
            feishu.fail_list_records = True
            controller = self._feishu_collection_controller(tmp, feishu, [bi_item])

            controller.start_feishu_collection({"table_id": "tbl_sop"})
            self._wait_for_state(controller, "completed")

            self.assertEqual(1, len(feishu.created_records))
            self.assertEqual(0, feishu.list_records_calls)
            self.assertFalse(any("飞书写入失败" in row.get("note", "") for row in controller.status()["logs"]))

    def test_account_protection_adds_longer_rest_after_every_five_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            sleeps = []
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=sleeps.append,
            )
            controller.save_upload(
                "input.xlsx",
                workbook_bytes([f"https://detail.tmall.com/item.htm?id={index}" for index in range(1, 7)]),
            )
            controller.save_settings({"intensity": "账号保护"})

            controller.start_collection()
            self._wait_for_state(controller, "completed")

            self.assertEqual(5, len(sleeps))
            self.assertGreaterEqual(sleeps[-1], 80)

    def test_ultra_conservative_adds_long_rest_after_five_real_fetches(self):
        with tempfile.TemporaryDirectory() as tmp:
            sleeps = []
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=sleeps.append,
            )
            controller.save_upload(
                "input.xlsx",
                workbook_bytes([f"https://detail.tmall.com/item.htm?id={index}" for index in range(1, 7)]),
            )
            controller.save_settings({"intensity": "超保守"})

            controller.start_collection()
            self._wait_for_state(controller, "completed")

            self.assertEqual(5, len(sleeps))
            self.assertGreaterEqual(min(sleeps), 45)
            self.assertLessEqual(max(sleeps), 390)
            self.assertGreaterEqual(sleeps[-1], 225)

    def _wait_for_state(self, controller, expected):
        deadline = time.time() + 5
        while time.time() < deadline:
            if controller.status()["run_state"] == expected:
                return
            time.sleep(0.02)
        self.fail(f"Timed out waiting for state {expected}; got {controller.status()['run_state']}")

    def _wait_for_calls(self, scraper, expected):
        deadline = time.time() + 5
        while time.time() < deadline:
            if len(scraper.calls) >= expected:
                return
            time.sleep(0.02)
        self.fail(f"Timed out waiting for {expected} calls; got {len(scraper.calls)}")

    def _feishu_collection_controller(self, tmp, feishu, items):
        controller = WebAppController(
            base_dir=Path(tmp),
            browser_manager=FakeBrowserManager(),
            scraper_factory=lambda: FakeScraper(),
            sleep=lambda seconds: None,
        )
        controller._feishu_client = feishu
        controller.state.feishu_connected = True
        controller.state.feishu_app_token = "base_token"
        controller.state.feishu_table_id = "tbl_sop"
        controller.bi_items = list(items)
        controller.state.bi_earphone_count = len(items)
        controller.state.link_count = len(items)
        controller.state.feishu_mapping = {
            "品牌": "品牌",
            "链接": "链接",
            "形态": "形态",
            "价格": "价格",
            "上架日期": "上架日期",
            "标题": "标题",
        }
        controller.state.feishu_field_types = {}
        controller.session = None
        if not items:
            controller.session = TaskSession(source_file=Path("."), link_rows=[], source_type="feishu", bi_items=[])
        return controller


class DashboardHttpTests(unittest.TestCase):
    def test_dashboard_home_and_status_endpoint(self):
        with tempfile.TemporaryDirectory() as tmp:
            controller = WebAppController(
                base_dir=Path(tmp),
                browser_manager=FakeBrowserManager(),
                scraper_factory=lambda: FakeScraper(),
                sleep=lambda seconds: None,
            )
            server = DashboardServer(("127.0.0.1", 0), DashboardRequestHandler, controller)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                base_url = f"http://127.0.0.1:{server.server_address[1]}"
                html = urllib.request.urlopen(base_url + "/", timeout=5).read().decode("utf-8")
                payload = urllib.request.urlopen(base_url + "/api/status", timeout=5).read()
                favicon_status = urllib.request.urlopen(base_url + "/favicon.ico", timeout=5).status
                status = json.loads(payload.decode("utf-8"))
            finally:
                server.shutdown()
                server.server_close()

            self.assertIn("淘天竞品监控工作台", html)
            self.assertIn("竞品监控", html)
            self.assertIn("飞书连接配置", html)
            self.assertIn("采集流程", html)
            self.assertIn("账号与验证", html)
            self.assertIn("采集日志", html)
            self.assertIn("采集设置", html)
            self.assertIn("超保守", html)
            self.assertIn("边界BI导入", html)
            self.assertIn('href="#top"', html)
            self.assertIn('href="#feishu-section"', html)
            self.assertIn('href="#bi-section"', html)
            self.assertIn('href="#verification-section"', html)
            self.assertIn('href="#logs-section"', html)
            self.assertIn("本地服务连接中断", html)
            self.assertIn("打开边界BI浏览器", html)
            self.assertIn("读取BI新品", html)
            self.assertIn("开始采集并写入飞书", html)
            self.assertIn("清空BI导入", html)
            self.assertIn("当前运行", html)
            self.assertEqual(200, favicon_status)
            self.assertIn("自动恢复飞书", html)
            self.assertIn("高级操作", html)
            self.assertNotIn("开始低频采集", html)
            self.assertNotIn("识别到 0 条链接", html)
            self.assertIn("我已完成验证，继续当前行", html)
            self.assertIn("跳过当前行", html)
            self.assertIn("复制本条", html)
            self.assertNotIn('class="nav-item"', html)
            self.assertEqual("idle", status["run_state"])
            self.assertIn("current_row", status)
            self.assertIn("can_resume", status)
            self.assertIn("bi_browser", status)
            self.assertIn("bi_import_count", status)
            self.assertIn("bi_earphone_count", status)
            self.assertIn("app_version", status)
            self.assertIn("app_path", status)
            self.assertIn("process_id", status)


if __name__ == "__main__":
    unittest.main()
