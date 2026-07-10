from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import os
from pathlib import Path
import random
import re
import sys
import time
import threading
import time
from typing import Any, Callable
import webbrowser

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .bi import BiBrowserSession, BiClient, BiImportConfig, BiTemplateItem
from .browser_session import LoginBrowserManager
from .cache import DailyResultCache
from .classifier import DEFAULT_SHAPE, HeuristicShapeClassifier, normalize_shape
from .deepseek import DeepSeekShapeClassifier
from .excel_io import ArchiveItem, LinkRow, RowResult, write_archive
from .feishu_client import FeishuClient, parse_app_token
from .feishu_config import (
    clear_feishu_credentials,
    load_feishu_credentials,
    save_feishu_credentials,
)
from .models import ProductSnapshot
from .progress import TaskProgressStore
from .safety import CrawlIntensity
from .scraper import PlaywrightProductScraper
from .settings import AppSettings, SettingsStore, app_data_dir
# template_filler removed in v8


VERIFICATION_STATUSES = {"needs_verification", "login_required", "captcha"}
APP_VERSION = "v8-feishu-direct"
DEFAULT_FEISHU_TABLE_NAME = "上新sop"
SAVED_FEISHU_SECRET_INVALID_MESSAGE = "保存的 App Secret 已失效，请重新填写后连接"


def _clean_feishu_app_id(value: Any) -> str:
    text = str(value or "").strip()
    if text == "cli_test":
        return ""
    matches = re.findall(r"(?=(cli_[A-Za-z0-9]+))", text)
    return matches[-1] if matches else text


def _clean_feishu_app_token(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text == "base_token":
        return ""
    url_match = re.search(r"https?://", text)
    if url_match:
        text = text[url_match.start():].strip()
    parsed = parse_app_token(text)
    return parsed or text


def _is_feishu_secret_invalid_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "app secret invalid" in text or "10014" in text


def _feishu_restore_error(exc: Exception) -> dict[str, Any]:
    error = (
        SAVED_FEISHU_SECRET_INVALID_MESSAGE
        if _is_feishu_secret_invalid_error(exc)
        else "飞书自动恢复失败，请重新连接飞书或检查已保存凭证"
    )
    return {
        "ok": False,
        "status": "restore_failed",
        "error": error,
        "detail": str(exc),
    }


def _select_feishu_table(tables: list[dict], preferred_table_id: str = "") -> dict | None:
    if not tables:
        return None
    preferred_table_id = str(preferred_table_id or "").strip()
    if preferred_table_id:
        selected = next((table for table in tables if str(table.get("id", "")) == preferred_table_id), None)
        if selected:
            return selected
    return next(
        (table for table in tables if str(table.get("name", "")) == DEFAULT_FEISHU_TABLE_NAME),
        tables[0],
    )


DASHBOARD_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>淘天竞品监控工作台</title>
  <style>
    :root {
      --ink: #172033;
      --muted: #6b7287;
      --soft: #f5f7fb;
      --paper: #ffffff;
      --line: #e4e8f1;
      --primary: #6252f3;
      --primary-dark: #3825c9;
      --green: #0f9f84;
      --green-dark: #05715e;
      --blue: #2675d8;
      --pink: #f06aa6;
      --amber: #b76b00;
      --red: #c33328;
      --feishu: #3370ff;
      --shadow: 0 20px 46px rgba(34, 45, 72, .08);
      --rail: #15172b;
    }
    * { box-sizing: border-box; }
    html { scroll-behavior: smooth; }
    body {
      margin: 0;
      color: var(--ink);
      background: var(--soft);
      font-family: "Microsoft YaHei UI", "Segoe UI", sans-serif;
    }
    .app-shell {
      display: grid;
      grid-template-columns: 248px minmax(0, 1fr);
      min-height: 100vh;
    }
    .rail {
      background: linear-gradient(180deg, #171935, #101225 72%, #171935);
      color: white;
      padding: 24px 18px;
      position: sticky;
      top: 0;
      min-height: 100vh;
    }
    .brand-lockup {
      display: grid;
      grid-template-columns: 42px 1fr;
      gap: 12px;
      align-items: center;
      margin-bottom: 28px;
    }
    .brand-mark {
      width: 42px;
      height: 42px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      background: linear-gradient(145deg, var(--primary), var(--pink));
      font-weight: 900;
    }
    .brand-title { font-weight: 900; line-height: 1.2; }
    .brand-sub { color: rgba(255,255,255,.58); font-size: 12px; margin-top: 2px; }
    .nav-list { display: grid; gap: 8px; }
    .nav-link {
      display: flex;
      align-items: center;
      gap: 10px;
      color: rgba(255,255,255,.72);
      padding: 11px 12px;
      border-radius: 8px;
      font-weight: 800;
      text-decoration: none;
    }
    .nav-link.active { color: white; background: rgba(255,255,255,.11); }
    .nav-link:hover, .nav-link:focus-visible { color: white; background: rgba(255,255,255,.12); }
    .nav-link:focus-visible { outline: 2px solid rgba(255,255,255,.86); outline-offset: 3px; }
    .nav-dot {
      width: 9px;
      height: 9px;
      border-radius: 50%;
      background: currentColor;
      opacity: .85;
    }
    .rail-panel {
      margin-top: 28px;
      padding: 14px;
      border: 1px solid rgba(255,255,255,.12);
      border-radius: 8px;
      background: rgba(255,255,255,.08);
      color: rgba(255,255,255,.78);
      font-size: 13px;
      line-height: 1.65;
    }
    .workspace {
      width: min(1480px, calc(100vw - 296px));
      margin: 0 auto;
      padding: 28px 24px 34px;
    }
    .topbar { display: flex; justify-content: space-between; align-items: flex-start; gap: 18px; margin-bottom: 18px; }
    .eyebrow { color: var(--primary); font-size: 13px; font-weight: 800; letter-spacing: .04em; }
    h1 { margin: 4px 0 6px; font-size: 34px; line-height: 1.15; }
    .sub { margin: 0; color: var(--muted); font-size: 15px; }
    .status-stack { display: flex; gap: 10px; flex-wrap: wrap; justify-content: flex-end; }
    .pill {
      background: rgba(255,255,255,.82);
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 10px 14px;
      color: var(--primary-dark);
      font-weight: 800;
      box-shadow: 0 8px 22px rgba(31, 66, 52, .06);
      white-space: nowrap;
    }
    .pill.warn { color: var(--amber); border-color: #f1d8a8; background: #fff8e9; }
    .pill.bad { color: var(--red); border-color: #f2c7c2; background: #fff2f0; }
    .metrics { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; margin-bottom: 16px; }
    .metric, .card {
      background: rgba(255,255,255,.94);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .metric { min-height: 112px; padding: 18px; position: relative; overflow: hidden; }
    .metric::after {
      content: "";
      position: absolute;
      right: 14px;
      top: 14px;
      width: 42px;
      height: 42px;
      border-radius: 50%;
      background: linear-gradient(145deg, rgba(98,82,243,.14), rgba(15,159,132,.14));
    }
    .metric .label, .field-label { color: var(--muted); font-size: 13px; font-weight: 700; }
    .metric .value { margin-top: 10px; font-size: 26px; line-height: 1.25; font-weight: 900; overflow-wrap: anywhere; }
    .metric .value.small { font-size: 16px; }
    .layout { display: grid; grid-template-columns: minmax(320px, 390px) minmax(0, 1fr); gap: 16px; align-items: start; }
    .side-stack { display: grid; gap: 14px; }
    .card { padding: 18px; }
    .card h2 { margin: 0 0 14px; font-size: 18px; }
    .workflow { display: grid; gap: 10px; }
    .step { display: grid; grid-template-columns: 32px 1fr; gap: 10px; align-items: start; padding: 12px; border: 1px solid var(--line); border-radius: 8px; background: #fbfcff; }
    .step-num { width: 32px; height: 32px; border-radius: 8px; display: grid; place-items: center; background: #eef0ff; color: var(--primary); font-weight: 900; }
    .step-title { font-weight: 900; }
    .step-note { color: var(--muted); font-size: 13px; margin-top: 4px; line-height: 1.45; }
    .verification-card { border-color: #bfe6df; background: linear-gradient(160deg, #ffffff, #f1fbf9); }
    .verification-text { color: var(--muted); line-height: 1.65; font-size: 14px; margin: 0 0 14px; }
    .field { display: grid; gap: 8px; margin-bottom: 12px; }
    input[type="password"], input[type="date"], input[type="number"], input[type="text"], select {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: white;
      padding: 12px;
      font: inherit;
    }
    .actions { display: flex; flex-wrap: wrap; gap: 9px; }
    button {
      border: 0;
      border-radius: 8px;
      padding: 11px 14px;
      font: inherit;
      font-weight: 900;
      cursor: pointer;
      background: var(--primary);
      color: white;
    }
    button.secondary { background: var(--green-dark); }
    button.blue { background: var(--blue); }
    button.warn { background: var(--amber); }
    button.quiet { color: var(--primary-dark); background: white; border: 1px solid var(--line); }
    button.feishu { background: var(--feishu); }
    button:disabled { opacity: .5; cursor: not-allowed; }
    .note { color: var(--muted); font-size: 13px; line-height: 1.65; margin: 10px 0 0; }
    .main-stack { display: grid; gap: 14px; }
    .log-card { min-height: 520px; }
    .log-head { display: flex; align-items: center; justify-content: space-between; gap: 12px; margin-bottom: 12px; }
    .log-head .actions { justify-content: flex-end; }
    .table-wrap { overflow-x: auto; border: 1px solid var(--line); border-radius: 8px; }
    table { width: 100%; border-collapse: collapse; font-size: 13px; background: white; }
    th, td { padding: 11px 10px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }
    th { color: var(--muted); font-weight: 900; background: #f7faf8; }
    td { overflow-wrap: anywhere; }
    tr:last-child td { border-bottom: 0; }
    .row-actions { display: flex; gap: 6px; flex-wrap: wrap; }
    .row-actions button { padding: 7px 9px; font-size: 12px; }
    .status-ok { color: var(--green); font-weight: 900; }
    .status-warn { color: var(--amber); font-weight: 900; }
    .status-bad { color: var(--red); font-weight: 900; }
    .toast { min-height: 22px; color: var(--muted); font-size: 13px; margin-top: 10px; }
    .settings-grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; }
    @media (max-width: 1100px) {
      .app-shell { grid-template-columns: 1fr; }
      .rail { position: relative; min-height: auto; padding: 16px; }
      .nav-list { grid-template-columns: repeat(5, minmax(0, 1fr)); }
      .nav-link { justify-content: center; font-size: 13px; }
      .rail-panel { display: none; }
      .workspace { width: min(100vw - 28px, 980px); padding-top: 18px; }
      .metrics, .layout, .settings-grid { grid-template-columns: 1fr; }
      .topbar { flex-direction: column; }
      .status-stack { justify-content: flex-start; }
      .log-head { align-items: flex-start; flex-direction: column; }
      .log-head .actions { width: 100%; justify-content: flex-start; }
      .log-head .actions button { flex: 1 1 140px; }
      h1 { font-size: 28px; }
    }
  </style>
</head>
<body>
  <div class="app-shell">
    <aside class="rail">
      <div class="brand-lockup">
        <div class="brand-mark">竞</div>
        <div>
          <div class="brand-title">竞品监控</div>
          <div class="brand-sub">BI · 淘宝 · 飞书</div>
        </div>
      </div>
      <nav class="nav-list" aria-label="主导航">
        <a class="nav-link active" href="#top"><span class="nav-dot"></span><span>工作台</span></a>
        <a class="nav-link" href="#feishu-section"><span class="nav-dot"></span><span>飞书</span></a>
        <a class="nav-link" href="#bi-section"><span class="nav-dot"></span><span>BI</span></a>
        <a class="nav-link" href="#verification-section"><span class="nav-dot"></span><span>验证</span></a>
        <a class="nav-link" href="#logs-section"><span class="nav-dot"></span><span>日志</span></a>
      </nav>
      <div class="rail-panel">
        <strong>当前策略</strong><br>
        默认使用超保守节奏；验证页出现后暂停等待人工处理。
      </div>
    </aside>
    <main class="workspace" id="top">
    <header class="topbar">
      <div>
        <div class="eyebrow">TAOTIAN COMPETITOR MONITOR</div>
        <h1>淘天竞品监控工作台</h1>
        <p class="sub">从边界BI读取耳机新品，复用淘宝登录态补全价格和形态，直连写入飞书多维表格。</p>
      </div>
      <div class="status-stack">
        <div class="pill" id="taskPill">任务状态：待命</div>
        <div class="pill" id="safetyPill">账号安全状态：正常</div>
      </div>
    </header>

    <section class="metrics">
      <div class="metric"><div class="label">BI新品</div><div class="value" id="linkCount">0</div></div>
      <div class="metric"><div class="label">采集进度</div><div class="value" id="progress">0 / 0</div></div>
      <div class="metric"><div class="label">风险状态</div><div class="value small" id="riskState">正常</div></div>
      <div class="metric"><div class="label">飞书连接</div><div class="value small" id="feishuConn">未连接</div></div>
    </section>

    <section class="layout">
      <div class="side-stack">
        <div class="card" id="feishu-section">
          <h2>采集流程</h2>
          <div class="workflow">
            <div class="step"><div class="step-num">1</div><div><div class="step-title">连接飞书</div><div class="step-note">填写飞书应用凭证和多维表格URL，选择目标数据表。</div></div></div>
            <div class="step"><div class="step-num">2</div><div><div class="step-title">登录边界BI</div><div class="step-note">打开边界BI浏览器，手动登录一次，登录态会保存在本机。</div></div></div>
            <div class="step"><div class="step-num">3</div><div><div class="step-title">读取BI新品</div><div class="step-note">选择统计范围和截止日期，只保留耳机相关新品。</div></div></div>
            <div class="step"><div class="step-num">4</div><div><div class="step-title">登录淘宝</div><div class="step-note">打开淘宝浏览器完成扫码或验证，用同一个浏览器采集详情页。</div></div></div>
            <div class="step"><div class="step-num">5</div><div><div class="step-title">写入飞书+归档</div><div class="step-note">按账号保护节奏采价，自动写入飞书多维表格，同时生成Excel归档副本。</div></div></div>
          </div>
        </div>

        <div class="card" id="bi-section">
          <h2>飞书连接配置</h2>
          <div class="field">
            <label class="field-label" for="feishuAppId">App ID</label>
            <input id="feishuAppId" type="text" placeholder="飞书应用 App ID">
          </div>
          <div class="field">
            <label class="field-label" for="feishuAppSecret">App Secret</label>
            <input id="feishuAppSecret" type="password" placeholder="飞书应用 App Secret">
          </div>
          <div class="field">
            <label class="field-label" for="feishuUrl">多维表格 URL</label>
            <input id="feishuUrl" type="text" placeholder="粘贴飞书多维表格链接或 app_token">
          </div>
          <div class="field">
            <label class="field-label" for="feishuTable">数据表</label>
            <select id="feishuTable"><option value="">正在恢复飞书连接...</option></select>
          </div>
          <div class="actions">
            <button class="feishu" id="feishuConnectBtn">保存并连接飞书</button>
            <button class="quiet" id="feishuRefreshFieldsBtn">刷新字段匹配</button>
          </div>
          <p class="note" id="feishuMappingMsg">字段匹配：等待连接...</p>
          <p class="note" id="feishuMsg">连接后自动按字段名匹配飞书表格。</p>
        </div>

        <div class="card">
          <h2>边界BI导入</h2>
          <div class="field">
            <label class="field-label" for="biDays">统计范围</label>
            <select id="biDays">
              <option value="7">7天</option>
              <option value="14">14天</option>
              <option value="30">30天</option>
            </select>
          </div>
          <div class="field">
            <label class="field-label" for="biEndDate">截止日期</label>
            <input id="biEndDate" type="date">
          </div>
          <div class="actions">
            <button class="secondary" id="openBiBtn">打开边界BI浏览器</button>
            <button class="blue" id="importBiBtn">读取BI新品</button>
            <button id="startFeishuImportBtn">开始采集并写入飞书</button>
            <button class="quiet" id="clearBiBtn">清空BI导入</button>
          </div>
          <p class="note" id="biMsg">先连接飞书选择数据表，再登录边界BI；工具只读取当前统计周期新品。</p>
        </div>

        <div class="card verification-card" id="verification-section">
          <h2>账号与验证</h2>
          <p class="verification-text" id="verificationText">当前无需处理。遇到扫码、短信或滑块验证时，工具会暂停并等待你人工完成。</p>
          <div class="actions">
            <button class="secondary" id="openLoginBtn">打开淘宝登录浏览器</button>
            <button class="blue" id="resumeBtn">我已完成验证，继续当前行</button>
            <button class="warn" id="skipBtn">跳过当前行</button>
          </div>
          <p class="note" id="loginMsg">滑块/短信/扫码验证需要你手动处理，工具不会自动破解。</p>
        </div>
      </div>

      <div class="main-stack">
        <div class="card log-card" id="logs-section">
          <div class="log-head">
            <h2>采集日志</h2>
            <div class="actions">
              <button class="quiet" id="copyLogsBtn">复制全部日志</button>
              <button class="quiet" id="copyLastErrorBtn">复制最后错误</button>
              <button class="quiet" id="exportLogsBtn">导出日志</button>
            </div>
          </div>
          <div class="table-wrap">
            <table>
              <thead><tr><th>时间</th><th>行号</th><th>状态</th><th>价格</th><th>耳机形态</th><th>原因</th><th>操作</th></tr></thead>
              <tbody id="logRows"><tr><td colspan="7">暂无日志</td></tr></tbody>
            </table>
          </div>
          <div class="toast" id="runMsg"></div>
        </div>

        <div class="settings-grid">
          <div class="card">
            <h2>采集设置</h2>
            <div class="field">
              <label class="field-label" for="intensity">采集强度</label>
              <select id="intensity">
                <option value="超保守">超保守：45-90 秒/条</option>
                <option value="账号保护">账号保护：20-40 秒/条</option>
                <option value="保守">保守：8-20 秒/条</option>
                <option value="标准">标准：5-12 秒/条</option>
                <option value="手动确认">手动确认：逐条处理</option>
              </select>
            </div>
            <p class="note">超保守每 5 个真实详情页访问后追加长休息，缓存命中不计入。</p>
          </div>
          <div class="card">
            <h2>DeepSeek 设置</h2>
            <div class="field">
              <label class="field-label" for="apiKey">API Key</label>
              <input id="apiKey" type="password" placeholder="DeepSeek API Key">
            </div>
            <button class="quiet" id="saveSettingsBtn">保存设置</button>
            <p class="note" id="settingsMsg">Key 只保存在本机设置中。</p>
          </div>
          <div class="card">
            <h2>归档副本</h2>
            <p class="note" id="outputInfo">采集完成后自动生成Excel归档副本。</p>
            <div class="actions">
              <button class="quiet" id="openOutputBtn">打开归档文件</button>
            </div>
          </div>
          <div class="card">
            <h2>高级操作</h2>
            <p class="note" id="runtimeInfo">当前运行：等待状态加载</p>
            <div class="actions" style="margin-bottom:10px">
              <button class="quiet" id="clearCacheBtn">清除缓存和进度</button>
            </div>
            <p class="note" id="cacheMsg">清除后，下次采集将重新抓取所有链接。</p>
            <details>
              <summary>浏览器控制</summary>
              <p class="note">这些操作通常不需要使用。关闭浏览器可能导致下次重新验证。</p>
              <div class="actions">
                <button class="quiet" id="closeLoginBtn">关闭淘宝浏览器</button>
                <button class="quiet" id="closeBiBtn">关闭边界BI浏览器</button>
              </div>
            </details>
          </div>
        </div>
      </div>
    </section>
    </main>
  </div>

  <script>
    const $ = (id) => document.getElementById(id);
    let latestStatus = null;
    const runStateText = {
      idle: "待命",
      running: "采集中",
      waiting_manual_verification: "等待人工验证",
      completed: "已完成",
      paused: "已暂停",
      failed: "异常"
    };
    const securityText = {
      normal: "正常",
      suggested_rest: "建议休息",
      needs_verification: "需要登录/验证",
      paused: "已暂停"
    };
    const logStatusText = {
      success: "采集成功",
      failed: "采集失败",
      error: "采集失败",
      needs_verification: "需要登录/验证",
      login_required: "需要登录/验证",
      captcha: "需要登录/验证",
      off_shelf: "商品下架/失效",
      skipped: "已跳过",
      bi_imported: "BI已读取",
      info: "提示"
    };
    const statusClass = (status) => {
      if (["success", "bi_imported", "info"].includes(status)) return "status-ok";
      if (["needs_verification", "login_required", "captcha", "off_shelf", "skipped"].includes(status)) return "status-warn";
      if (status) return "status-bad";
      return "";
    };
    async function api(path, options = {}) {
      try {
        const res = await fetch(path, options);
        const data = await res.json();
        if (!res.ok || data.ok === false) throw new Error(data.error || "请求失败");
        return data;
      } catch (err) {
        if (err instanceof TypeError && /fetch/i.test(err.message || "")) {
          throw new Error("本地服务连接中断：请确认工作台 exe 仍在运行；如果刚关闭或重启过页面，请重新打开新版 exe 生成的工作台页面。");
        }
        throw err;
      }
    }
    function setMsg(id, text, bad = false) {
      $(id).textContent = text || "";
      $(id).className = "note" + (bad ? " status-bad" : "");
    }
    function formatRow(row) {
      return [
        "时间：" + (row.time || ""),
        "行号：" + (row.row || ""),
        "状态：" + (logStatusText[row.status] || row.status || ""),
        "价格：" + (row.price || ""),
        "耳机形态：" + (row.shape || ""),
        "原因：" + (row.note || ""),
        "链接：" + (row.url || ""),
        "下一步建议：" + (row.next_action || "")
      ].join("\n");
    }
    function formatLogRows(rows) {
      if (!rows || !rows.length) return "暂无日志";
      return rows.map(formatRow).join("\n\n");
    }
    async function copyText(text, message) {
      try {
        await navigator.clipboard.writeText(text);
      } catch (err) {
        const box = document.createElement("textarea");
        box.value = text;
        box.style.position = "fixed";
        box.style.left = "-1000px";
        document.body.appendChild(box);
        box.focus();
        box.select();
        document.execCommand("copy");
        document.body.removeChild(box);
      }
      setMsg("runMsg", message);
    }
    function fullLogText() {
      const s = latestStatus || {};
      return [
        "任务状态：" + (runStateText[s.run_state] || s.run_state || ""),
        "账号安全状态：" + (securityText[s.security_state] || s.security_state || ""),
        "BI新品数量：" + (s.bi_import_count || 0),
        "BI耳机新品：" + (s.bi_earphone_count || 0),
        "BI导入状态：" + (s.bi_import_status || ""),
        "飞书状态：" + (s.feishu_connected ? "已连接" : "未连接"),
        "飞书表格：" + (s.feishu_table_name || ""),
        "字段匹配：" + (s.feishu_mapping_summary || ""),
        "当前行：" + (s.current_row || ""),
        "当前链接：" + (s.current_url || ""),
        "暂停原因：" + (s.pause_reason || ""),
        "下一步建议：" + (s.next_action || ""),
        "结果文件：" + (s.output_path || ""),
        "",
        formatLogRows(s.logs || [])
      ].join("\n");
    }
    function lastErrorText() {
      const rows = ((latestStatus || {}).logs || []).slice().reverse();
      const row = rows.find((item) => item.status && item.status !== "success") || rows[0];
      return row ? formatRow(row) : "暂无错误日志";
    }
    function renderLogs(rows) {
      const displayRows = rows.slice().reverse();
      $("logRows").innerHTML = displayRows.length ? displayRows.map((row, index) => `
        <tr>
          <td>${row.time || ""}</td>
          <td>${row.row || ""}</td>
          <td class="${statusClass(row.status)}">${logStatusText[row.status] || row.status || ""}</td>
          <td>${row.price || ""}</td>
          <td>${row.shape || ""}</td>
          <td>${row.note || ""}</td>
          <td><div class="row-actions">
            <button class="quiet" data-copy-log-index="${index}">复制本条</button>
            ${row.row && ["failed", "error"].includes(row.status) ? '<button class="quiet" data-retry-row="' + row.row + '">重试本条</button>' : ""}
            ${latestStatus && latestStatus.can_resume && latestStatus.current_row === row.row ? '<button class="quiet" data-skip-row="' + row.row + '">跳过当前行</button>' : ""}
          </div></td>
        </tr>`).join("") : '<tr><td colspan="7">暂无日志</td></tr>';
    }
    async function refreshStatus() {
      try {
        const s = await api("/api/status");
        latestStatus = s;
        $("linkCount").textContent = s.bi_earphone_count || s.link_count || 0;
        $("progress").textContent = (s.completed || 0) + " / " + (s.link_count || 0);
        $("riskState").textContent = securityText[s.security_state] || s.security_state || "正常";
        $("runtimeInfo").textContent = "当前运行：" + (s.app_version || "") + " | 进程 " + (s.process_id || "") + " | " + (s.app_path || "");
        $("taskPill").textContent = "任务状态：" + (runStateText[s.run_state] || s.run_state);
        $("safetyPill").textContent = "账号安全状态：" + (securityText[s.security_state] || s.security_state);
        $("taskPill").className = "pill " + (s.run_state === "waiting_manual_verification" ? "warn" : "");
        $("safetyPill").className = "pill " + (s.security_state === "normal" ? "" : "warn");
        $("intensity").value = s.intensity || "超保守";
        if (!$("apiKey").value && s.deepseek_api_key) $("apiKey").value = s.deepseek_api_key;
        if (!$("feishuAppId").value && s.feishu_app_id_saved) $("feishuAppId").value = s.feishu_app_id_saved;
        if (!$("feishuAppSecret").value && s.feishu_secret_saved) $("feishuAppSecret").placeholder = "已保存 · 无需重新填写";
        if (!$("feishuUrl").value && s.feishu_url_saved) $("feishuUrl").value = s.feishu_url_saved;
        $("resumeBtn").disabled = !s.can_resume;
        $("skipBtn").disabled = !s.current_row;
        $("verificationText").textContent = s.current_row
          ? "第 " + s.current_row + " 行需要处理：" + (s.pause_reason || "请检查当前链接") + "。" + (s.next_action || "")
          : "当前无需处理。遇到扫码、短信或滑块验证时，请在已打开的淘宝浏览器中人工完成，然后点击继续当前行。";
        $("outputInfo").textContent = s.output_path ? "归档文件：" + s.output_path : "采集完成后自动生成Excel归档副本。";
        if (s.bi_import_message) {
          $("biMsg").textContent = s.bi_import_message + "；耳机新品 " + (s.bi_earphone_count || 0) + " 条，受检品牌 " + (s.checked_brand_count || 0) + " 个";
        }
        $("feishuConn").textContent = s.feishu_connected ? "已连接" : "未连接";
        $("feishuConn").style.color = s.feishu_connected ? "var(--green)" : "var(--muted)";
        $("feishuMappingMsg").textContent = s.feishu_mapping_summary || "字段匹配：等待连接...";
        $("feishuMappingMsg").style.color = s.feishu_connected ? "var(--green)" : "var(--muted)";
        $("startFeishuImportBtn").disabled = !(s.feishu_connected && s.bi_earphone_count > 0 && s.run_state !== "running");
        if ($("feishuTable") && document.activeElement !== $("feishuTable")) {
          if (s.feishu_tables && s.feishu_tables.length) {
            $("feishuTable").innerHTML = s.feishu_tables.map(t => '<option value="' + t.id + '"' + (t.id === s.feishu_table_id ? ' selected' : '') + '>' + t.name + '</option>').join("");
          }
        }
        renderLogs(s.logs || []);
        if (s.run_state !== "idle") setMsg("runMsg", "任务状态：" + (runStateText[s.run_state] || s.run_state) + (s.pause_reason ? "，" + s.pause_reason : ""));
      } catch (err) {
        setMsg("runMsg", err.message, true);
      }
    }
    async function restoreFeishuDefaults() {
      $("feishuTable").innerHTML = '<option value="">正在恢复飞书连接...</option>';
      setMsg("feishuMsg", "正在恢复飞书连接...");
      try {
        const data = await api("/api/feishu/restore", { method: "POST" });
        if (data.status === "restored") {
          setMsg("feishuMsg", data.message || "已自动恢复飞书连接");
        } else if (data.status === "empty") {
          $("feishuTable").innerHTML = '<option value="">请先连接飞书</option>';
          setMsg("feishuMsg", data.message || "首次使用请填写飞书配置");
        }
      } catch (err) {
        setMsg("feishuMsg", "自动恢复飞书失败：" + err.message, true);
      }
    }
    // Feishu connect
    $("feishuConnectBtn").onclick = async () => {
      try {
        const data = await api("/api/feishu/connect", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({
            app_id: $("feishuAppId").value,
            app_secret: $("feishuAppSecret").value,
            app_token: $("feishuUrl").value
          })
        });
        setMsg("feishuMsg", data.message || "飞书连接成功");
        $("feishuConn").textContent = "已连接";
        $("feishuConn").style.color = "var(--green)";
        $("feishuTable").innerHTML = (data.tables || []).map(t => '<option value="' + t.id + '"' + (t.id === data.table_id ? ' selected' : '') + '>' + t.name + '</option>').join("");
        $("feishuMappingMsg").textContent = data.mapping_summary || "";
        refreshStatus();
      } catch (err) { setMsg("feishuMsg", err.message, true); }
    };
    $("feishuRefreshFieldsBtn").onclick = async () => {
      const tableId = $("feishuTable").value;
      if (!tableId) return setMsg("feishuMsg", "请先连接飞书并选择数据表", true);
      try {
        const data = await api("/api/feishu/fields", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ table_id: tableId })
        });
        $("feishuMappingMsg").textContent = data.mapping_summary || "字段匹配完成";
        setMsg("feishuMsg", data.message || "");
        refreshStatus();
      } catch (err) { setMsg("feishuMsg", err.message, true); }
    };
    $("feishuTable").onchange = () => $("feishuRefreshFieldsBtn").click();
    // BI
    $("openBiBtn").onclick = async () => {
      try {
        const data = await api("/api/open-bi-browser", { method: "POST" });
        setMsg("biMsg", data.message || "已打开边界BI浏览器");
        refreshStatus();
      } catch (err) { setMsg("biMsg", err.message, true); }
    };
    $("closeBiBtn").onclick = async () => {
      try {
        await api("/api/close-bi-browser", { method: "POST" });
        setMsg("biMsg", "边界BI浏览器已关闭");
        refreshStatus();
      } catch (err) { setMsg("biMsg", err.message, true); }
    };
    $("importBiBtn").onclick = async () => {
      try {
        const data = await api("/api/import-bi-goods", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ days_type: Number($("biDays").value || 7), end_date: $("biEndDate").value })
        });
        setMsg("biMsg", data.message || "BI新品读取完成");
        refreshStatus();
      } catch (err) { setMsg("biMsg", err.message, true); }
    };
    $("clearBiBtn").onclick = async () => {
      try {
        const data = await api("/api/clear-bi-import", { method: "POST" });
        setMsg("biMsg", data.message || "已清空BI导入结果");
        refreshStatus();
      } catch (err) { setMsg("biMsg", err.message, true); }
    };
    $("startFeishuImportBtn").onclick = async () => {
      const tableId = $("feishuTable").value;
      if (!tableId) return setMsg("feishuMsg", "请先连接飞书并选择数据表", true);
      try {
        const data = await api("/api/start-feishu-import", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ table_id: tableId })
        });
        setMsg("runMsg", data.message);
        refreshStatus();
      } catch (err) { setMsg("runMsg", err.message, true); }
    };
    // Login
    $("openLoginBtn").onclick = async () => {
      try {
        const data = await api("/api/open-login-browser", { method: "POST" });
        setMsg("loginMsg", data.message || "已打开淘宝登录浏览器");
        refreshStatus();
      } catch (err) { setMsg("loginMsg", err.message, true); }
    };
    $("closeLoginBtn").onclick = async () => {
      try {
        await api("/api/close-login-browser", { method: "POST" });
        setMsg("loginMsg", "淘宝浏览器已关闭");
        refreshStatus();
      } catch (err) { setMsg("loginMsg", err.message, true); }
    };
    // Settings
    $("saveSettingsBtn").onclick = async () => {
      try {
        await api("/api/settings", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ deepseek_api_key: $("apiKey").value, intensity: $("intensity").value })
        });
        $("apiKey").value = "";
        setMsg("settingsMsg", "设置已保存到本机");
        refreshStatus();
      } catch (err) { setMsg("settingsMsg", err.message, true); }
    };
    // Actions
    $("resumeBtn").onclick = async () => {
      try {
        const data = await api("/api/resume-current-row", { method: "POST" });
        setMsg("runMsg", data.message);
        refreshStatus();
      } catch (err) { setMsg("runMsg", err.message, true); }
    };
    $("skipBtn").onclick = async () => {
      try {
        const data = await api("/api/skip-current-row", { method: "POST" });
        setMsg("runMsg", data.message);
        refreshStatus();
      } catch (err) { setMsg("runMsg", err.message, true); }
    };
    $("openOutputBtn").onclick = async () => {
      try {
        const data = await api("/api/open-output", { method: "POST" });
        setMsg("runMsg", data.message);
      } catch (err) { setMsg("runMsg", err.message, true); }
    };
    $("exportLogsBtn").onclick = async () => {
      try {
        const data = await api("/api/export-logs", { method: "POST" });
        setMsg("runMsg", data.message);
      } catch (err) { setMsg("runMsg", err.message, true); }
    };
    $("copyLogsBtn").onclick = () => copyText(fullLogText(), "全部日志已复制");
    $("copyLastErrorBtn").onclick = () => copyText(lastErrorText(), "最后错误已复制");
    $("logRows").onclick = async (event) => {
      const target = event.target;
      if (!target || !target.dataset) return;
      if (target.dataset.copyLogIndex) {
        const displayRows = (latestStatus.logs || []).slice().reverse();
        return copyText(formatRow(displayRows[Number(target.dataset.copyLogIndex)] || {}), "本条日志已复制");
      }
      const row = Number(target.dataset.retryRow || target.dataset.skipRow || 0);
      if (!row) return;
      if (target.dataset.retryRow) {
        const data = await api("/api/retry-row", { method: "POST", headers: { "Content-Type": "application/json" }, body: JSON.stringify({ row }) });
        setMsg("runMsg", data.message);
        refreshStatus();
      }
      if (target.dataset.skipRow) {
        const data = await api("/api/skip-current-row", { method: "POST", headers: { "Content-Type": "application/json" }, body: JSON.stringify({ row }) });
        setMsg("runMsg", data.message);
        refreshStatus();
      }
    };
    $("clearCacheBtn").onclick = async () => {
      if (!confirm("确认清除所有缓存和采集进度？\n清除后下次采集将重新抓取所有链接。")) return;
      try {
        const data = await api("/api/clear-cache", { method: "POST" });
        setMsg("cacheMsg", data.message);
        refreshStatus();
      } catch (err) { setMsg("cacheMsg", err.message, true); }
    };
    $("biEndDate").value = new Date().toISOString().slice(0, 10);
    restoreFeishuDefaults().finally(refreshStatus);
    setInterval(refreshStatus, 1200);
  </script>
</body>
</html>
"""



@dataclass
class RuntimeState:
    selected_file: Path | None = None
    selected_file_name: str = ""
    link_count: int = 0
    output_path: Path | None = None
    run_state: str = "idle"
    security_state: str = "normal"
    completed: int = 0
    failed: int = 0
    reason: str = ""
    current_row: int | None = None
    current_url: str = ""
    current_index: int = 0
    pending_rows: int = 0
    pause_reason: str = ""
    next_action: str = ""
    can_resume: bool = False
    logs: list[dict[str, Any]] = field(default_factory=list)
    bi_import_count: int = 0
    bi_earphone_count: int = 0
    bi_import_status: str = "idle"
    bi_import_message: str = ""
    feishu_connected: bool = False
    feishu_app_token: str = ""
    feishu_table_id: str = ""
    feishu_table_name: str = ""
    feishu_tables: list[dict] = field(default_factory=list)
    feishu_field_count: int = 0
    feishu_field_types: dict = field(default_factory=dict)
    feishu_mapping_summary: str = ""
    feishu_mapping: dict = field(default_factory=dict)
    checked_brands: list[str] = field(default_factory=list)  # v8: 所有受检品牌
    checked_brand_count: int = 0
    _keepalive_last: float = 0.0  # v8: 上次会话保活时间戳

@dataclass
class TaskSession:
    source_file: Path
    link_rows: list[LinkRow]
    results: list[RowResult] = field(default_factory=list)
    current_index: int = 0
    source_type: str = "excel"
    bi_items: list[BiTemplateItem] = field(default_factory=list)


class WebAppController:
    def __init__(
        self,
        *,
        base_dir: Path | None = None,
        settings_store: SettingsStore | None = None,
        browser_manager: LoginBrowserManager | None = None,
        bi_browser_manager: BiBrowserSession | None = None,
        bi_client: BiClient | None = None,
        scraper_factory: Callable[[], Any] | None = None,
        folder_picker: Callable[[], str | None] | None = None,
        sleep: Callable[[float], None] | None = None,
    ) -> None:
        self.base_dir = Path(base_dir) if base_dir else app_data_dir()
        self.upload_dir = self.base_dir / "uploads"
        self.runs_dir = self.base_dir / "runs"
        self.browser_profile_dir = self.base_dir / "browser-profile"
        self.bi_browser_profile_dir = self.base_dir / "bi-browser-profile"
        self.settings_store = settings_store or SettingsStore(self.base_dir / "settings.json")
        self.settings = self.settings_store.load()
        self.browser_manager = browser_manager or LoginBrowserManager(user_data_dir=self.browser_profile_dir)
        self.bi_browser_manager = bi_browser_manager or BiBrowserSession(user_data_dir=self.bi_browser_profile_dir)
        self.bi_client = bi_client or BiClient(session=self.bi_browser_manager)
        self.scraper_factory = scraper_factory or (
            lambda: PlaywrightProductScraper(browser_session=self.browser_manager)
        )
        self.folder_picker = folder_picker or self._select_folder_dialog
        self.sleep = sleep or time.sleep
        self.state = RuntimeState()
        self.session: TaskSession | None = None
        self.bi_items: list[BiTemplateItem] = []
        self.cache = DailyResultCache(self.base_dir / "cache" / "daily_results.json")
        self.progress = TaskProgressStore(self.base_dir / "progress" / "task_progress.json")
        self._lock = threading.RLock()
        self._run_thread: threading.Thread | None = None
        self._last_heartbeat: float = time.time()
        self._feishu_client: FeishuClient | None = None
        self._feishu_creds = load_feishu_credentials()
        if self._feishu_creds:
            try:
                app_id = _clean_feishu_app_id(self._feishu_creds.get("app_id", ""))
                app_token = _clean_feishu_app_token(self._feishu_creds.get("app_token", ""))
                self._feishu_creds = {
                    **self._feishu_creds,
                    "app_id": app_id,
                    "app_token": app_token,
                }
                if not app_id or not self._feishu_creds.get("app_secret"):
                    raise ValueError("invalid saved feishu credentials")
                self._feishu_client = FeishuClient(
                    app_id,
                    self._feishu_creds["app_secret"],
                )
                self.state.feishu_connected = True
                self.state.feishu_app_token = app_token
                self.state.feishu_table_id = self._feishu_creds.get("table_id", "")
                self.state.feishu_table_name = self._feishu_creds.get("table_name", "")
            except Exception:
                self._feishu_client = None

    def status(self) -> dict[str, Any]:
        self._last_heartbeat = time.time()
        with self._lock:
            output_dir, output_dir_source = self._output_dir_status_locked()
            return {
                "ok": True,
                "app_version": APP_VERSION,
                "app_path": str(self._running_app_path()),
                "process_id": os.getpid(),
                "selected_file_name": self.state.selected_file_name,
                "link_count": self.state.link_count,
                "output_path": str(self.state.output_path) if self.state.output_path else "",
                "output_dir": output_dir,
                "output_dir_source": output_dir_source,
                "run_state": self.state.run_state,
                "security_state": self.state.security_state,
                "completed": self.state.completed,
                "failed": self.state.failed,
                "reason": self.state.reason,
                "current_row": self.state.current_row,
                "current_url": self.state.current_url,
                "current_index": self.state.current_index,
                "pending_rows": self.state.pending_rows,
                "pause_reason": self.state.pause_reason,
                "next_action": self.state.next_action,
                "can_resume": self.state.can_resume,
                "logs": list(self.state.logs[-300:]),
                "deepseek_configured": bool(self.settings.deepseek_api_key),
                "intensity": self.settings.intensity,
                "active_mode": "feishu" if self.bi_items else ("excel" if self.session else "idle"),
                "can_start_feishu_import": bool(self.state.feishu_connected and self.bi_items),
                "can_start_collection": bool(self.state.feishu_connected and self.bi_items),
                "can_start_bi_fill": bool(self.bi_items and self.state.selected_file),
                "browser_profile_dir": str(self.browser_profile_dir),
                "login_browser": self.browser_manager.status(),
                "bi_browser": self.bi_browser_manager.status(),
                "bi_import_count": self.state.bi_import_count,
                "bi_earphone_count": self.state.bi_earphone_count,
                "bi_import_status": self.state.bi_import_status,
                "bi_import_message": self.state.bi_import_message,
                "feishu_connected": self.state.feishu_connected,
                "feishu_app_token": self.state.feishu_app_token,
                "feishu_table_id": self.state.feishu_table_id,
                "feishu_table_name": self.state.feishu_table_name,
                "feishu_tables": self.state.feishu_tables,
                "feishu_field_count": self.state.feishu_field_count,
                "feishu_mapping_summary": self.state.feishu_mapping_summary,
                "feishu_mapping": self.state.feishu_mapping,
                "checked_brands": list(self.state.checked_brands),
                "checked_brand_count": self.state.checked_brand_count,
                "deepseek_api_key": self.settings.deepseek_api_key,
                "feishu_app_id_saved": self._feishu_creds.get("app_id", "") if self._feishu_creds else "",
                "feishu_secret_saved": bool(self._feishu_creds and self._feishu_creds.get("app_secret")),
                "feishu_url_saved": self._feishu_creds.get("app_token", "") if self._feishu_creds else "",
                "feishu_table_id_saved": self._feishu_creds.get("table_id", "") if self._feishu_creds else "",
                "feishu_table_name_saved": self._feishu_creds.get("table_name", "") if self._feishu_creds else "",
            }


    def save_upload(self, filename: str, content: bytes) -> dict[str, Any]:
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        safe_name = self._safe_filename(filename)
        path = self.upload_dir / safe_name
        path.write_bytes(content)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            link_rows: list[LinkRow] = []
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=3).value
                text = str(value or "").strip()
                if "taobao.com" in text or "tmall.com" in text:
                    link_rows.append(LinkRow(row_number=row, url=text))
        finally:
            wb.close()

        with self._lock:
            self.state.selected_file = path
            self.state.selected_file_name = safe_name
            self.state.link_count = len(link_rows)
            self.state.pending_rows = len(link_rows)
            self.state.completed = 0
            self.state.failed = 0
            self.state.current_row = None
            self.state.current_url = ""
            self.state.output_path = None
            self.session = TaskSession(source_file=path, link_rows=link_rows, source_type="excel")
        return {"ok": True, "file_name": safe_name, "path": str(path), "link_count": len(link_rows)}

    def save_settings(self, payload: dict[str, Any]) -> dict[str, Any]:
        api_key = str(payload.get("deepseek_api_key") or self.settings.deepseek_api_key).strip()
        intensity = str(payload.get("intensity") or self.settings.intensity).strip()
        if intensity not in {"超保守", "账号保护", "保守", "标准", "手动确认"}:
            raise ValueError("采集强度只能选择：超保守、账号保护、保守、标准、手动确认")
        output_dir = self.settings.output_dir
        if "output_dir" in payload:
            output_dir = self._normalize_output_dir(str(payload.get("output_dir") or ""))
        self.settings = AppSettings(deepseek_api_key=api_key, intensity=intensity, output_dir=output_dir)
        self.settings_store.save(self.settings)
        with self._lock:
            if self.state.feishu_connected and self._feishu_client:
                save_feishu_credentials(
                    self._feishu_client.app_id,
                    self._feishu_client.app_secret,
                    self.state.feishu_app_token,
                    table_id=self.state.feishu_table_id,
                    table_name=self.state.feishu_table_name,
                )
        return {"ok": True, "message": "设置已保存", "output_dir": output_dir}

    def select_output_dir(self) -> dict[str, Any]:
        selected = self.folder_picker()
        if not selected:
            output_dir, output_dir_source = self._output_dir_status_locked()
            return {
                "ok": True,
                "status": "cancelled",
                "message": "未选择新的保存文件夹",
                "output_dir": output_dir,
                "output_dir_source": output_dir_source,
            }
        output_dir = self._normalize_output_dir(selected)
        self.settings = AppSettings(
            deepseek_api_key=self.settings.deepseek_api_key,
            intensity=self.settings.intensity,
            output_dir=output_dir,
        )
        self.settings_store.save(self.settings)
        return {
            "ok": True,
            "status": "selected",
            "message": "结果保存位置已更新",
            "output_dir": output_dir,
            "output_dir_source": "自定义",
        }

    def _normalize_output_dir(self, value: str) -> str:
        text = str(value or "").strip().strip('"')
        if not text:
            return ""
        path = Path(text).expanduser()
        if not path.exists() or not path.is_dir():
            raise ValueError(f"保存目录不存在或不可用：{path}")
        return str(path)

    def _output_dir_status_locked(self) -> tuple[str, str]:
        if self.settings.output_dir:
            path = Path(self.settings.output_dir)
            if path.exists() and path.is_dir():
                return str(path), "自定义"
            return str(path), "无效"
        if self.state.selected_file:
            return str(self.state.selected_file.parent), "模板目录"
        return str(self.runs_dir), "默认目录"

    def _resolve_output_dir_for_write(self) -> Path:
        if self.settings.output_dir:
            return Path(self._normalize_output_dir(self.settings.output_dir))
        if self.state.selected_file:
            return self.state.selected_file.parent
        return self.runs_dir

    @staticmethod
    def _select_folder_dialog() -> str | None:
        try:
            import tkinter as tk
            from tkinter import filedialog

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            selected = filedialog.askdirectory(title="选择结果保存文件夹")
            root.destroy()
            return selected or None
        except Exception as exc:
            raise RuntimeError(f"无法打开文件夹选择窗口：{exc}") from exc

    def open_login_browser(self) -> dict[str, Any]:
        result = self.browser_manager.open()
        return {"ok": True, **result}

    def close_login_browser(self) -> dict[str, Any]:
        self.browser_manager.close()
        return {"ok": True, "message": "淘宝浏览器已关闭"}

    def open_bi_browser(self) -> dict[str, Any]:
        result = self.bi_browser_manager.open()
        return {"ok": True, **result}

    def close_bi_browser(self) -> dict[str, Any]:
        self.bi_browser_manager.close()
        return {"ok": True, "message": "边界BI浏览器已关闭"}

    # ---- Feishu ----

    def feishu_connect(self, payload: dict[str, Any]) -> dict[str, Any]:
        app_id = _clean_feishu_app_id(payload.get("app_id", ""))
        app_secret = str(payload.get("app_secret", "")).strip()
        app_token_input = _clean_feishu_app_token(payload.get("app_token", ""))
        saved_table_id = ""
        saved_table_name = ""
        # v9.1: 若未填凭据，自动从本地加密存储加载
        if not app_secret or not app_id or not app_token_input:
            saved = load_feishu_credentials()
            if saved:
                app_id = app_id or _clean_feishu_app_id(saved.get("app_id", ""))
                app_secret = app_secret or str(saved.get("app_secret", "")).strip()
                app_token_input = app_token_input or _clean_feishu_app_token(saved.get("app_token", ""))
                saved_table_id = str(saved.get("table_id", "")).strip()
                saved_table_name = str(saved.get("table_name", "")).strip()
        if not app_id or not app_secret:
            raise ValueError("请填写飞书 App ID 和 App Secret")
        app_token = app_token_input
        try:
            client = FeishuClient(app_id, app_secret)
            spreadsheets, _ = client.list_spreadsheets()
        except Exception as exc:
            if _is_feishu_secret_invalid_error(exc):
                raise ValueError("App Secret 无效，请重新填写后连接飞书") from exc
            raise
        with self._lock:
            previous_client = self._feishu_client
            previous_creds = dict(self._feishu_creds) if self._feishu_creds else None
            previous_feishu_state = {
                "connected": self.state.feishu_connected,
                "app_token": self.state.feishu_app_token,
                "table_id": self.state.feishu_table_id,
                "table_name": self.state.feishu_table_name,
                "tables": list(self.state.feishu_tables),
                "field_count": self.state.feishu_field_count,
                "field_types": dict(self.state.feishu_field_types),
                "mapping_summary": self.state.feishu_mapping_summary,
                "mapping": dict(self.state.feishu_mapping),
            }
        with self._lock:
            self._feishu_client = client
            self._feishu_creds = {
                "app_id": app_id,
                "app_secret": app_secret,
                "app_token": app_token_input,
                "table_id": saved_table_id or self.state.feishu_table_id,
                "table_name": saved_table_name or self.state.feishu_table_name,
            }
            self.state.feishu_connected = True
            self.state.feishu_app_token = app_token
            self.state.feishu_tables = []
        if app_token:
            try:
                return self._load_feishu_tables_and_fields(
                    app_token,
                    select_table_id=saved_table_id or self.state.feishu_table_id,
                )
            except Exception:
                with self._lock:
                    self._feishu_client = previous_client
                    self._feishu_creds = previous_creds
                    self.state.feishu_connected = previous_feishu_state["connected"]
                    self.state.feishu_app_token = previous_feishu_state["app_token"]
                    self.state.feishu_table_id = previous_feishu_state["table_id"]
                    self.state.feishu_table_name = previous_feishu_state["table_name"]
                    self.state.feishu_tables = previous_feishu_state["tables"]
                    self.state.feishu_field_count = previous_feishu_state["field_count"]
                    self.state.feishu_field_types = previous_feishu_state["field_types"]
                    self.state.feishu_mapping_summary = previous_feishu_state["mapping_summary"]
                    self.state.feishu_mapping = previous_feishu_state["mapping"]
                raise
        with self._lock:
            save_feishu_credentials(
                app_id,
                app_secret,
                app_token_input,
                table_id=self.state.feishu_table_id,
                table_name=self.state.feishu_table_name,
            )
        tables_summary = ", ".join(s["name"] for s in spreadsheets[:5])
        if len(spreadsheets) > 5:
            tables_summary += f" 等共{len(spreadsheets)}个"
        return {
            "ok": True,
            "status": "connected",
            "message": f"飞书连接成功，找到 {len(spreadsheets)} 个多维表格，请粘贴目标表格URL",
            "spreadsheet_count": len(spreadsheets),
            "tables": [],
            "table_id": self.state.feishu_table_id,
            "table_name": self.state.feishu_table_name,
            "mapping_summary": "",
        }

    def feishu_load_fields(self, payload: dict[str, Any]) -> dict[str, Any]:
        table_id = str(payload.get("table_id", "")).strip()
        if not table_id:
            raise ValueError("请提供数据表 ID")
        with self._lock:
            app_token = self.state.feishu_app_token
            client = self._feishu_client
        if not app_token or not client:
            raise ValueError("请先连接飞书")
        return self._load_feishu_tables_and_fields(app_token, select_table_id=table_id)

    def feishu_restore(self) -> dict[str, Any]:
        saved = load_feishu_credentials()
        if not saved:
            return {"ok": True, "status": "empty", "message": "没有已保存的飞书配置"}
        app_id = _clean_feishu_app_id(saved.get("app_id", ""))
        app_secret = str(saved.get("app_secret", "")).strip()
        app_token_input = _clean_feishu_app_token(saved.get("app_token", ""))
        table_id = str(saved.get("table_id", "")).strip()
        if not app_id or not app_secret:
            return {"ok": False, "status": "invalid", "error": "已保存的飞书配置不完整"}
        app_token = app_token_input
        try:
            client = FeishuClient(app_id, app_secret)
        except Exception as exc:
            with self._lock:
                self.state.feishu_connected = False
            return _feishu_restore_error(exc)
        with self._lock:
            self._feishu_client = client
            self._feishu_creds = {
                **saved,
                "app_id": app_id,
                "app_token": app_token_input,
            }
            self.state.feishu_connected = True
            self.state.feishu_app_token = app_token
            self.state.feishu_table_id = table_id
            self.state.feishu_table_name = str(saved.get("table_name", "")).strip()
        if app_token:
            try:
                restored = self._load_feishu_tables_and_fields(app_token, select_table_id=table_id)
            except Exception as exc:
                with self._lock:
                    self.state.feishu_connected = False
                    self.state.feishu_mapping_summary = ""
                return _feishu_restore_error(exc)
            restored["status"] = "restored"
            restored["message"] = f"已自动恢复飞书连接，默认数据表：{self.state.feishu_table_name}"
            return restored
        return {"ok": True, "status": "restored", "message": "已恢复飞书凭证，请选择多维表格"}

    def _load_feishu_tables_and_fields(self, app_token: str, select_table_id: str = "") -> dict[str, Any]:
        client = self._feishu_client
        if not client:
            raise ValueError("飞书未连接")
        tables = client.list_tables(app_token)
        mapping_summary = ""
        field_types = {}
        mapping = {}
        fields_list = []
        selected_table = None
        if tables:
            selected_table = _select_feishu_table(tables, select_table_id)
            target_id = selected_table["id"]
            fields_list = client.list_fields(app_token, target_id)
            field_names = {f["name"] for f in fields_list}
            field_types = {f["name"]: f["type"] for f in fields_list}
            expected = ["品牌", "链接", "形态", "价格", "上架日期", "标题"]
            matched = [k for k in expected if k in field_names]
            if matched:
                mapping = {k: k for k in matched}
                mapping_summary = "✅ 自动匹配 " + str(len(matched)) + "/" + str(len(expected)) + " 字段: " + ", ".join(matched)
            else:
                mapping_summary = "⚠ 未匹配，飞书字段: " + ", ".join(list(field_names)[:6])
        with self._lock:
            self.state.feishu_table_id = (selected_table or {}).get("id", "") if tables else ""
            self.state.feishu_table_name = (selected_table or {}).get("name", "") if tables else ""
            self.state.feishu_tables = tables
            self.state.feishu_field_count = len(fields_list)
            self.state.feishu_field_types = field_types
            self.state.feishu_mapping_summary = mapping_summary
            self.state.feishu_mapping = mapping
            self.state.feishu_app_token = app_token
            if self._feishu_client:
                self._feishu_creds = {
                    "app_id": self._feishu_client.app_id,
                    "app_secret": self._feishu_client.app_secret,
                    "app_token": app_token,
                    "table_id": self.state.feishu_table_id,
                    "table_name": self.state.feishu_table_name,
                }
                save_feishu_credentials(
                    self._feishu_client.app_id,
                    self._feishu_client.app_secret,
                    app_token,
                    table_id=self.state.feishu_table_id,
                    table_name=self.state.feishu_table_name,
                )
        return {
            "ok": True,
            "tables": tables,
            "fields": fields_list,
            "field_types": field_types,
            "table_id": self.state.feishu_table_id,
            "table_name": self.state.feishu_table_name,
            "mapping_summary": mapping_summary,
            "mapping": mapping,
            "message": (
                f"已加载 {len(tables)} 个数据表，默认数据表：{self.state.feishu_table_name}"
                if self.state.feishu_table_name
                else f"已加载 {len(tables)} 个数据表"
            ),
        }

    def import_bi_goods(self, payload: dict[str, Any]) -> dict[str, Any]:
        config = BiImportConfig.from_payload(payload)
        result = self.bi_client.import_goods(config, template_brands=None)  # v8: no template needed
        with self._lock:
            self.bi_items = list(result.items)
            self.state.bi_import_count = result.import_count
            self.state.bi_earphone_count = result.earphone_count
            self.state.bi_import_status = "completed"
            self.state.bi_import_message = result.message
            self.state.link_count = result.earphone_count
            self.state.pending_rows = result.earphone_count
            self.state.checked_brands = list(result.checked_brands)
            self.state.checked_brand_count = len(result.checked_brands)
            self._append_log(
                {
                    "status": "bi_imported",
                    "price": "",
                    "shape": "",
                    "note": result.message,
                    "next_action": "确认飞书已连接后，点击开始采集并写入飞书",
                }
            )
        return {
            "ok": True,
            "status": "completed",
            "bi_import_count": result.import_count,
            "bi_earphone_count": result.earphone_count,
            "message": result.message,
        }

    def clear_bi_import(self) -> dict[str, Any]:
        with self._lock:
            if self.state.run_state in {"running", "waiting_manual_verification"}:
                raise RuntimeError("当前任务正在进行，不能清空 BI 导入结果")
            excel_session = self.session if self.session and self.session.source_type == "excel" else None
            self.bi_items = []
            self.session = excel_session
            self.state.bi_import_count = 0
            self.state.bi_earphone_count = 0
            self.state.bi_import_status = "idle"
            self.state.bi_import_message = ""
            self.state.link_count = len(excel_session.link_rows) if excel_session else 0
            self.state.pending_rows = len(excel_session.link_rows) if excel_session else 0
            self.state.completed = 0
            self.state.failed = 0
            self.state.current_row = None
            self.state.current_url = ""
            self.state.current_index = 0
            self.state.pause_reason = ""
            self.state.next_action = ""
            self.state.can_resume = False
            self.state.output_path = None
            self._append_log(
                {
                    "status": "info",
                    "price": "",
                    "shape": "",
                    "note": "已清空当前 BI 导入结果，请重新读取 BI 新品",
                    "next_action": "点击读取BI新品获取最新页面记录",
                }
            )
            return {
                "ok": True,
                "status": "cleared",
                "message": "已清空当前 BI 导入结果",
                "link_count": self.state.link_count,
            }

    def start_feishu_collection(self, payload: dict[str, Any]) -> dict[str, Any]:
        table_id = str(payload.get("table_id", "")).strip()
        if not table_id:
            raise ValueError("请提供目标数据表 ID")
        with self._lock:
            self.state.feishu_table_id = table_id
            if self.state.run_state == "running":
                raise RuntimeError("当前已有采集任务正在运行")
            if not self.bi_items:
                raise ValueError("请先读取 BI 新品")
            if not self.state.feishu_connected or not self._feishu_client:
                raise ValueError("请先连接飞书并选择数据表")
            if not self.state.feishu_app_token:
                raise ValueError("请先输入飞书多维表格 URL")
            # Refresh fields for the selected table
            self._load_feishu_tables_and_fields(self.state.feishu_app_token, select_table_id=table_id)
            link_rows = [
                LinkRow(row_number=index + 1, url=item.fetch_link or item.goods_link)
                for index, item in enumerate(self.bi_items)
            ]
            self.session = TaskSession(
                source_file=Path("."),
                link_rows=link_rows,
                source_type="feishu",
                bi_items=list(self.bi_items),
            )
            self.state.link_count = len(link_rows)
            self.state.completed = 0
            self.state.failed = 0
            self.state.pending_rows = len(link_rows)
            self._set_running_state()
            if not link_rows:
                self._finish_collection_locked()
                return {"ok": True, "status": "completed", "message": "未筛出耳机新品，无需写入"}
        self._start_worker()
        return {"ok": True, "status": "started", "message": "已开始采集 BI 新品并写入飞书"}



    def resume_current_row(self) -> dict[str, Any]:
        self._ensure_ready_to_run()
        with self._lock:
            if self.state.run_state != "waiting_manual_verification" or not self.state.can_resume:
                raise RuntimeError("当前没有等待人工验证的行")
            self._append_log(
                {
                    "row": self.state.current_row,
                    "url": self.state.current_url,
                    "status": "info",
                    "price": "",
                    "shape": "",
                    "note": "复用淘宝浏览器会话，继续当前商品",
                    "next_action": "",
                }
            )
            # v9: 验证解决后强制刷新会话，避免短时间内再次弹验证
            self.state._keepalive_last = 0.0
            self._set_running_state()
        self._start_worker()
        return {"ok": True, "status": "resumed", "message": "已继续当前行采集"}

    def skip_current_row(self, row: int | None = None) -> dict[str, Any]:
        with self._lock:
            if not self.session:
                raise ValueError("请先读取 BI 新品")
            target_row = row or self.state.current_row
            if not target_row:
                raise RuntimeError("当前没有可跳过的行")
            index = self._index_for_row(target_row)
            if index is None:
                raise ValueError("没有找到指定行号")
            self.session.current_index = index
            link_row = self.session.link_rows[index]
            result = RowResult(
                row_number=link_row.row_number,
                price="获取失败",
                shape=DEFAULT_SHAPE,
                status="failed",
                note="用户跳过当前行",
            )
            self._record_row_result(link_row, result, count_completed=False, count_failed=True)
            self.session.current_index = index + 1
            self._set_running_state()
        self._start_worker()
        return {"ok": True, "status": "skipped", "message": "已跳过当前行，继续处理后续链接"}

    def retry_row(self, row: int) -> dict[str, Any]:
        self._ensure_ready_to_run()
        with self._lock:
            index = self._index_for_row(row)
            if index is None:
                raise ValueError("没有找到指定行号")
            self.session.current_index = index  # type: ignore[union-attr]
            self._set_running_state()
        self._start_worker(stop_after_one=True)
        return {"ok": True, "status": "retrying", "message": f"已开始重试第 {row} 行"}

    def export_logs(self) -> dict[str, Any]:
        with self._lock:
            self.runs_dir.mkdir(parents=True, exist_ok=True)
            path = self.runs_dir / f"采集日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            lines = [
                "\t".join(
                    [
                        str(row.get("time", "")),
                        str(row.get("row", "")),
                        self._display_status(str(row.get("status", ""))),
                        str(row.get("price", "")),
                        str(row.get("shape", "")),
                        str(row.get("note", "")),
                        str(row.get("url", "")),
                    ]
                )
                for row in self.state.logs
            ]
        path.write_text("\n".join(lines), encoding="utf-8")
        return {"ok": True, "path": str(path), "message": f"日志已导出：{path}"}

    def open_output(self) -> dict[str, Any]:
        with self._lock:
            output_path = self.state.output_path
        if not output_path or not output_path.exists():
            raise ValueError("当前还没有可打开的结果文件")
        if os.name == "nt":
            os.startfile(output_path)  # type: ignore[attr-defined]
        else:
            webbrowser.open(output_path.as_uri())
        return {"ok": True, "message": "已打开结果文件"}

    def clear_cache(self) -> dict[str, Any]:
        with self._lock:
            if self.state.run_state == "running":
                return {"ok": False, "error": "采集运行中，无法清除缓存"}
            count = self.cache.clear()
            self.progress.clear()
            return {"ok": True, "message": f"已清除 {count} 条缓存和采集进度"}

    def start_collection(self) -> dict[str, Any]:
        self._ensure_ready_to_run()
        with self._lock:
            if self.session and self.session.current_index >= len(self.session.link_rows):
                self.session.current_index = 0
            self._set_running_state()
        self._start_worker()
        return {"ok": True, "status": "started", "message": "采集已启动，工具会按当前强度低频处理链接"}

    def start_bi_template_fill(self) -> dict[str, Any]:
        with self._lock:
            if self.state.run_state == "running":
                raise RuntimeError("当前已有采集任务正在运行")
            if not self.state.selected_file:
                raise ValueError("请先上传需要回填的 Excel 模板")
            link_rows = [
                LinkRow(row_number=index + 1, url=item.fetch_link or item.goods_link)
                for index, item in enumerate(self.bi_items)
            ]
            self.session = TaskSession(
                source_file=self.state.selected_file,
                link_rows=link_rows,
                source_type="bi_template",
                bi_items=list(self.bi_items),
            )
            self.state.link_count = len(link_rows)
            self.state.completed = 0
            self.state.failed = 0
            self.state.pending_rows = len(link_rows)
            self._set_running_state()
            if not link_rows:
                self._finish_collection_locked()
                return {"ok": True, "status": "completed", "message": "未筛出耳机新品，模板保持空白并输出副本"}
        self._start_worker()
        return {"ok": True, "status": "started", "message": "已开始补全 BI 新品价格并回填模板"}

    def _ensure_ready_to_run(self) -> None:
        with self._lock:
            if self.state.run_state == "running":
                raise RuntimeError("当前已有采集任务正在运行")
            if not self.session or not self.session.link_rows:
                raise ValueError("请先读取 BI 新品")

    def _set_running_state(self) -> None:
        self.state.run_state = "running"
        self.state.security_state = "normal"
        self.state.pause_reason = ""
        self.state.next_action = ""
        self.state.can_resume = False

    def _start_worker(self, *, stop_after_one: bool = False) -> None:
        self._run_thread = threading.Thread(
            target=self._run_collection_loop,
            kwargs={"stop_after_one": stop_after_one},
            name="taotian-collection",
            daemon=True,
        )
        self._run_thread.start()

    def _run_collection_loop(self, *, stop_after_one: bool = False) -> None:
        try:
            scraper = self.scraper_factory()
            classifier = DeepSeekShapeClassifier(api_key=self.settings.deepseek_api_key)
            processed = 0
            real_fetches = 0
            while True:
                with self._lock:
                    if not self.session:
                        return
                    if self.session.current_index >= len(self.session.link_rows):
                        self._finish_collection_locked()
                        return
                    link_row = self.session.link_rows[self.session.current_index]
                    self.state.current_row = link_row.row_number
                    self.state.current_url = link_row.url
                    self.state.current_index = self.session.current_index
                    self.state.pending_rows = len(self.session.link_rows) - self.session.current_index

                cached = self.cache.get(link_row.url)
                if cached:
                    cached_status = str(cached.get("status", "success"))
                    row_result = RowResult(
                        row_number=link_row.row_number,
                        price=str(cached.get("price", "")),
                        shape="" if cached_status == "off_shelf" else normalize_shape(str(cached.get("shape", DEFAULT_SHAPE))),
                        status=cached_status,
                        note="使用当日缓存，未重复访问",
                    )
                    with self._lock:
                        self._record_row_result(link_row, row_result, count_completed=True, count_failed=False)
                        self.session.current_index += 1  # type: ignore[union-attr]
                    processed += 1
                    if stop_after_one:
                        with self._lock:
                            self._finish_collection_locked()
                        return
                    continue

                snapshot = scraper.fetch(link_row.url)
                real_fetches += 1
                if snapshot.status in VERIFICATION_STATUSES:
                    with self._lock:
                        self.state.run_state = "waiting_manual_verification"
                        self.state.security_state = "needs_verification"
                        self.state.pause_reason = snapshot.note or "页面进入登录/验证，需要人工处理"
                        self.state.next_action = "请在已打开的淘宝浏览器完成登录/验证后继续当前行"
                        self.state.can_resume = True
                        self._append_or_update_verification_log(
                            {
                                "row": link_row.row_number,
                                "url": link_row.url,
                                "status": snapshot.status,
                                "price": "获取失败",
                                "shape": "",
                                "note": self.state.pause_reason,
                                "next_action": self.state.next_action,
                            }
                        )
                    return

                row_result = self._snapshot_to_result(link_row, snapshot, classifier)
                with self._lock:
                    processed_status = row_result.status in {"success", "off_shelf"}
                    self._record_row_result(
                        link_row,
                        row_result,
                        count_completed=processed_status,
                        count_failed=not processed_status,
                    )
                    self.session.current_index += 1  # type: ignore[union-attr]
                processed += 1
                if row_result.status in {"success", "off_shelf"}:
                    self.cache.set(
                        link_row.url,
                        {
                            "price": row_result.price,
                            "shape": row_result.shape,
                            "status": row_result.status,
                            "title": snapshot.title,
                        },
                    )
                if stop_after_one:
                    with self._lock:
                        self._finish_collection_locked()
                    return
                with self._lock:
                    should_sleep = self.session.current_index < len(self.session.link_rows)  # type: ignore[union-attr]
                if should_sleep and self.sleep is not None:
                    delay = self._selected_intensity().next_delay_seconds()
                    if self.settings.intensity == "超保守" and real_fetches > 0 and real_fetches % 5 == 0:
                        delay += random.randint(180, 300)
                    elif self.settings.intensity == "账号保护" and real_fetches > 0 and real_fetches % 5 == 0:
                        delay += random.randint(60, 120)
                    self.sleep(delay)
        except Exception as exc:
            with self._lock:
                self.state.run_state = "failed"
                self.state.security_state = "paused"
                self.state.reason = str(exc)
                self._append_log({"status": "failed", "note": str(exc), "next_action": "复制错误并检查配置"})

    def _snapshot_to_result(
        self,
        link_row: LinkRow,
        snapshot: ProductSnapshot,
        classifier: HeuristicShapeClassifier,
    ) -> RowResult:
        if snapshot.status == "success":
            shape = classifier.classify(snapshot.title, snapshot.detail_text)
            return RowResult(
                row_number=link_row.row_number,
                price=snapshot.price,
                shape=normalize_shape(shape.shape),
                status="success",
                note=snapshot.note or shape.reason,
            )
        if snapshot.status == "off_shelf":
            return RowResult(
                row_number=link_row.row_number,
                price="商品下架/失效",
                shape="",
                status="off_shelf",
                note=snapshot.note or "详情页提示商品已下架",
            )
        return RowResult(
            row_number=link_row.row_number,
            price="获取失败",
            shape=DEFAULT_SHAPE,
            status="failed",
            note=snapshot.note or "页面未成功采集，按默认入耳式填写",
        )

    def _record_row_result(
        self,
        link_row: LinkRow,
        result: RowResult,
        *,
        count_completed: bool,
        count_failed: bool,
    ) -> None:
        if self.session:
            self.session.results.append(result)
        if count_completed:
            self.state.completed += 1
        if count_failed:
            self.state.failed += 1
        self.progress.record_row(
            source_file=str(self.state.selected_file or ""),
            row=link_row.row_number,
            url=link_row.url,
            status=result.status,
            price=result.price,
            shape=result.shape,
            note=result.note,
        )
        self._append_log(
            {
                "row": link_row.row_number,
                "url": link_row.url,
                "status": result.status,
                "price": result.price,
                "shape": result.shape,
                "note": result.note,
                "next_action": "" if result.status in {"success", "off_shelf"} else "可重试本条或跳过",
            }
        )

    def _finish_collection_locked(self) -> None:
        if not self.session:
            self.state.run_state = "completed"
            return
        # --- 飞书直写 ---
        feishu_count = 0
        if self._feishu_client and self.state.feishu_app_token and self.state.feishu_table_id:
            try:
                records = self._build_feishu_records(self.session)
                if records:
                    feishu_count = self._feishu_client.batch_create_records(
                        self.state.feishu_app_token,
                        self.state.feishu_table_id,
                        records,
                        self.state.feishu_field_types,
                    )
                self._append_log({
                    "status": "info",
                    "price": "",
                    "shape": "",
                    "note": f"已写入飞书多维表格 {feishu_count} 条记录",
                    "next_action": "",
                })
            except Exception as exc:
                self._append_log({
                    "status": "failed",
                    "price": "",
                    "shape": "",
                    "note": f"飞书写入失败: {exc}",
                    "next_action": "请检查飞书连接和字段配置后重试",
                })
        # --- Excel 归档副本 ---
        archive_items = self._build_archive_items(self.session)
        output_dir = self._resolve_output_dir_for_write()
        try:
            if self.session.source_type == "excel" and self.session.source_file.exists():
                output_path = self._write_excel_result_copy(self.session, output_dir)
            elif self.session.source_type == "bi_template" and self.session.source_file.exists():
                output_path = self._write_bi_template_result_copy(self.session, output_dir)
            else:
                output_path = write_archive(archive_items, output_dir=output_dir)
            self.state.output_path = output_path
            self._append_log({
                "status": "info",
                "price": "",
                "shape": "",
                "note": f"归档副本已生成: {output_path}",
                "next_action": "",
            })
        except Exception as exc:
            self._append_log({
                "status": "failed",
                "price": "",
                "shape": "",
                "note": f"归档副本生成失败: {exc}",
                "next_action": "",
            })
        self.state.run_state = "completed"
        self.state.security_state = "normal"
        self.state.pause_reason = ""
        self.state.next_action = ""
        self.state.can_resume = False
        self.state.current_row = None
        self.state.current_url = ""
        self.state.pending_rows = 0

    def _write_excel_result_copy(self, session: TaskSession, output_dir: Path) -> Path:
        wb = load_workbook(session.source_file)
        try:
            ws = wb.active
            for result in session.results:
                if result.status == "success":
                    self._write_row_cells(ws, result.row_number, {4: result.price, 5: result.shape})
                elif result.status == "off_shelf":
                    self._write_row_cells(ws, result.row_number, {4: "", 5: ""})
            output_path = self._result_copy_path(session.source_file, output_dir)
            wb.save(output_path)
            return output_path
        finally:
            wb.close()

    def _write_bi_template_result_copy(self, session: TaskSession, output_dir: Path) -> Path:
        wb = load_workbook(session.source_file)
        try:
            ws = wb.active
            is_new_template = self._looks_like_new_template(ws)
            result_by_row = {result.row_number: result for result in session.results}
            brand_starts = self._brand_start_rows(ws)
            brand_cursor: dict[str, int] = {}
            written_brands: set[str] = set()

            for index, item in enumerate(session.bi_items, start=1):
                result = result_by_row.get(index)
                if not result:
                    continue
                if result.status == "off_shelf":
                    self._append_log({
                        "row": result.row_number,
                        "url": item.goods_link,
                        "status": "skipped",
                        "price": "",
                        "shape": "",
                        "note": "商品下架/失效，已跳过，不回填模板",
                        "next_action": "",
                    })
                    continue

                start_row = brand_starts.get(item.brand, ws.max_row + 1)
                target_row = brand_cursor.get(item.brand, start_row)
                brand_cursor[item.brand] = target_row + 1
                written_brands.add(item.brand)
                if target_row > ws.max_row:
                    ws.cell(row=target_row, column=1).value = item.brand

                date_text = item.display_date
                if is_new_template and date_text and not date_text.endswith("号"):
                    date_text += "号"
                if is_new_template:
                    values = {
                        1: item.brand,
                        2: date_text,
                        3: item.goods_link,
                        4: item.goods_name,
                        5: result.shape,
                        6: result.price,
                    }
                else:
                    values = {
                        1: item.brand,
                        2: date_text,
                        3: item.goods_link,
                        4: result.shape,
                        5: result.price,
                    }
                self._write_row_cells(ws, target_row, values)

            if not is_new_template:
                for brand, start_row in brand_starts.items():
                    if brand not in written_brands:
                        self._write_row_cells(ws, start_row, {2: "无上新"})

            output_path = self._result_copy_path(session.source_file, output_dir)
            wb.save(output_path)
            return output_path
        finally:
            wb.close()

    def _result_copy_path(self, source_file: Path, output_dir: Path) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = source_file.suffix or ".xlsx"
        return output_dir / f"{source_file.stem}_已补全_{timestamp}{suffix}"

    def _write_row_cells(self, ws: Any, row: int, values: dict[int, Any]) -> None:
        if not values:
            return
        self._unmerge_intersecting(ws, row, min(values), max(values))
        for column, value in values.items():
            ws.cell(row=row, column=column).value = value

    @staticmethod
    def _unmerge_intersecting(ws: Any, row: int, min_column: int, max_column: int) -> None:
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= row <= max_row and max(min_col, min_column) <= min(max_col, max_column):
                ws.unmerge_cells(str(merged_range))

    @staticmethod
    def _looks_like_new_template(ws: Any) -> bool:
        headers = [str(ws.cell(row=2, column=column).value or "").strip() for column in range(1, 7)]
        return headers[:6] == ["品牌", "上架日期", "链接", "标题", "形态", "价格"]

    @staticmethod
    def _brand_start_rows(ws: Any) -> dict[str, int]:
        starts: dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            brand = str(ws.cell(row=row, column=1).value or "").strip()
            if brand and brand not in starts and brand not in {"品牌", "店铺", "品牌名称"}:
                starts[brand] = row
        return starts

    def _build_feishu_records(self, session: TaskSession) -> list[dict]:
        """将采集结果构建为飞书批量写入的记录列表，字段名按动态匹配
        有上新的品牌写入完整行，无上新的品牌仅写入品牌名。"""
        mapping = self.state.feishu_mapping
        result_by_row = {result.row_number: result for result in session.results}

        sorted_items = sorted(enumerate(session.bi_items, start=1), key=lambda x: (x[1].brand, x[1].on_sale_time))
        earphone_brands = {item.brand for _, item in sorted_items if item.brand}

        records: list[dict] = []
        written_brands: set[str] = set()
        for index, item in sorted_items:
            result = result_by_row.get(index)
            if not result:
                continue
            if result.status == "off_shelf":
                self._append_log({
                    "row": result.row_number,
                    "status": "skipped",
                    "price": "",
                    "shape": "",
                    "url": item.goods_link,
                    "note": "商品下架/失效，已跳过，不写入飞书",
                    "next_action": "",
                })
                continue
            record: dict[str, Any] = {}
            if "品牌" in mapping and item.brand:
                record["品牌"] = item.brand
                written_brands.add(item.brand)
            if "链接" in mapping and item.goods_link:
                record["链接"] = item.goods_link
            if "标题" in mapping and item.goods_name:
                record["标题"] = item.goods_name
            if "形态" in mapping and result.shape:
                record["形态"] = result.shape
            if "价格" in mapping and result.price:
                record["价格"] = result.price
            if "上架日期" in mapping and item.on_sale_time:
                record["上架日期"] = item.on_sale_time
            if record:
                records.append(record)

        # v8: 为无上新的受检品牌补写空行（按品牌名排序）
        no_product_brands = sorted(
            [b for b in self.state.checked_brands if b not in earphone_brands]
        )
        for brand in no_product_brands:
            if "品牌" in mapping:
                records.append({"品牌": brand})
                self._append_log({
                    "status": "info",
                    "price": "",
                    "shape": "",
                    "note": f"品牌 {brand} 本周期无新品，仅写入品牌名",
                    "next_action": "",
                })
        return records

    def _build_archive_items(self, session: TaskSession) -> list[ArchiveItem]:
        """从采集结果构建归档数据，含无上新品牌空行。v9: 按品牌排序"""
        result_by_row = {result.row_number: result for result in session.results}
        items: list[ArchiveItem] = []
        written_brands: set[str] = set()

        # v9: 按品牌排序
        sorted_items = sorted(enumerate(session.bi_items, start=1), key=lambda x: (x[1].brand, x[1].on_sale_time))

        for index, item in sorted_items:
            result = result_by_row.get(index)
            if not result:
                continue
            written_brands.add(item.brand)
            items.append(ArchiveItem(
                row=index,
                brand=item.brand,
                on_sale_date=item.on_sale_time,
                goods_link=item.goods_link,
                goods_name=item.goods_name,
                shape=result.shape,
                price=result.price,
                status=result.status,
                note=result.note,
            ))
        # v8: 无上新品牌也写入归档（v9: 按品牌名排序）
        no_product_brands = sorted(
            [b for b in self.state.checked_brands if b not in written_brands]
        )
        for brand in no_product_brands:
            items.append(ArchiveItem(
                row=len(items) + 1,
                brand=brand,
                on_sale_date="",
                goods_link="",
                goods_name="",
                shape="",
                price="",
                status="no_new_products",
                note="本周期无新品",
            ))
        return items

    def _append_log(self, event: dict[str, Any]) -> None:
        row = dict(event)
        row["time"] = datetime.now().strftime("%H:%M:%S")
        self.state.logs.append(row)
        if len(self.state.logs) > 300:
            self.state.logs = self.state.logs[-300:]

    def _append_or_update_verification_log(self, event: dict[str, Any]) -> None:
        for existing in reversed(self.state.logs):
            if existing.get("row") != event.get("row"):
                break
            if existing.get("status") in VERIFICATION_STATUSES:
                existing.update(event)
                existing["time"] = datetime.now().strftime("%H:%M:%S")
                return
        self._append_log(event)

    def _index_for_row(self, row: int) -> int | None:
        if not self.session:
            return None
        for index, link_row in enumerate(self.session.link_rows):
            if link_row.row_number == row:
                return index
        return None

    def _selected_intensity(self) -> CrawlIntensity:
        if self.settings.intensity == "超保守":
            return CrawlIntensity.ultra_conservative()
        if self.settings.intensity == "账号保护":
            return CrawlIntensity.account_protection()
        if self.settings.intensity == "标准":
            return CrawlIntensity.standard()
        if self.settings.intensity == "手动确认":
            return CrawlIntensity.manual_confirmation()
        return CrawlIntensity.conservative()

    @staticmethod
    def _display_status(status: str) -> str:
        return {
            "success": "采集成功",
            "failed": "采集失败",
            "error": "采集失败",
            "needs_verification": "需要登录/验证",
            "login_required": "需要登录/验证",
            "captcha": "需要登录/验证",
            "off_shelf": "商品下架/失效",
            "skipped": "已跳过",
            "bi_imported": "BI已读取",
            "info": "提示",
        }.get(status, status)

    @staticmethod
    def _running_app_path() -> Path:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve()
        return Path(__file__).resolve()

    @staticmethod
    def _safe_filename(filename: str) -> str:
        name = Path(filename).name or "upload.xlsx"
        return re.sub(r"[^A-Za-z0-9._\-\u4e00-\u9fff（）()【】 ]+", "_", name)


class DashboardServer(ThreadingHTTPServer):
    allow_reuse_address = True

    def __init__(
        self,
        server_address: tuple[str, int],
        handler_class: type[BaseHTTPRequestHandler],
        controller: WebAppController,
    ) -> None:
        super().__init__(server_address, handler_class)
        self.controller = controller


class DashboardRequestHandler(BaseHTTPRequestHandler):
    server: DashboardServer

    def log_message(self, format: str, *args: Any) -> None:
        return

    def do_GET(self) -> None:
        if self.path in {"/", "/index.html"}:
            self._send_bytes(DASHBOARD_HTML.encode("utf-8"), content_type="text/html; charset=utf-8")
            return
        if self.path == "/favicon.ico":
            self._send_bytes(b"", content_type="image/x-icon")
            return
        if self.path == "/api/status":
            self._send_json(self.server.controller.status())
            return
        self._send_json({"ok": False, "error": "未找到接口"}, HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        try:
            if self.path == "/api/settings":
                self._send_json(self.server.controller.save_settings(self._read_json()))
            elif self.path == "/api/select-output-dir":
                self._send_json(self.server.controller.select_output_dir())
            elif self.path == "/api/open-login-browser":
                self._send_json(self.server.controller.open_login_browser())
            elif self.path == "/api/close-login-browser":
                self._send_json(self.server.controller.close_login_browser())
            elif self.path == "/api/open-bi-browser":
                self._send_json(self.server.controller.open_bi_browser())
            elif self.path == "/api/close-bi-browser":
                self._send_json(self.server.controller.close_bi_browser())
            elif self.path == "/api/import-bi-goods":
                self._send_json(self.server.controller.import_bi_goods(self._read_json()))
            elif self.path == "/api/clear-bi-import":
                self._send_json(self.server.controller.clear_bi_import())
            elif self.path == "/api/feishu/connect":
                self._send_json(self.server.controller.feishu_connect(self._read_json()))
            elif self.path == "/api/feishu/restore":
                self._send_json(self.server.controller.feishu_restore())
            elif self.path == "/api/feishu/fields":
                self._send_json(self.server.controller.feishu_load_fields(self._read_json()))
            elif self.path == "/api/start-feishu-import":
                self._send_json(self.server.controller.start_feishu_collection(self._read_json()))
            elif self.path == "/api/resume-current-row":
                self._send_json(self.server.controller.resume_current_row())
            elif self.path == "/api/skip-current-row":
                payload = self._read_json()
                row = int(payload["row"]) if payload.get("row") else None
                self._send_json(self.server.controller.skip_current_row(row=row))
            elif self.path == "/api/retry-row":
                payload = self._read_json()
                self._send_json(self.server.controller.retry_row(int(payload["row"])))
            elif self.path == "/api/export-logs":
                self._send_json(self.server.controller.export_logs())
            elif self.path == "/api/open-output":
                self._send_json(self.server.controller.open_output())
            elif self.path == "/api/clear-cache":
                self._send_json(self.server.controller.clear_cache())
            else:
                self._send_json({"ok": False, "error": "未找到接口"}, HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self._send_json({"ok": False, "error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def _read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))


    def _send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        self._send_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            status=status,
            content_type="application/json; charset=utf-8",
        )

    def _send_bytes(
        self,
        content: bytes,
        *,
        status: HTTPStatus = HTTPStatus.OK,
        content_type: str = "application/octet-stream",
    ) -> None:
        self.send_response(status.value)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(content)


def find_available_port(start: int = 17888, attempts: int = 20) -> int:
    import socket

    for port in range(start, start + attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            try:
                sock.bind(("127.0.0.1", port))
            except OSError:
                continue
            return port
    raise RuntimeError("没有找到可用的本地端口")


def _should_shutdown_for_stale_heartbeat(
    controller: WebAppController,
    *,
    now: float,
    timeout: float,
) -> bool:
    with controller._lock:
        protected_states = {"running", "waiting_manual_verification", "completed", "failed", "paused"}
        if controller.state.run_state in protected_states:
            return False
        return now - controller._last_heartbeat > timeout


def _heartbeat_watchdog(
    server: DashboardServer,
    controller: WebAppController,
    timeout: float = 10.0,
) -> None:
    """后台守护线程：检测前端心跳，网页关闭后自动退出服务器。"""
    while True:
        time.sleep(3)
        if _should_shutdown_for_stale_heartbeat(controller, now=time.time(), timeout=timeout):
            print("网页已关闭，自动退出。")
            server.shutdown()
            break


def run_server(*, port: int | None = None, open_browser: bool = True) -> None:
    selected_port = port or find_available_port()
    controller = WebAppController()
    server = DashboardServer(("127.0.0.1", selected_port), DashboardRequestHandler, controller)
    url = f"http://127.0.0.1:{selected_port}"
    if open_browser:
        webbrowser.open(url)
    print(f"淘天竞品监控已启动：{url}")
    watchdog = threading.Thread(
        target=_heartbeat_watchdog,
        args=(server, controller),
        daemon=True,
    )
    watchdog.start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        controller.close_login_browser()
        controller.close_bi_browser()
        server.server_close()


def main() -> None:
    run_server()
